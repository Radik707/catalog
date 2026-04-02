"""
make_photo_sheet.py — Создание Excel-файла для ручного сопоставления фото с товарами

Использование:
    python make_photo_sheet.py
    python make_photo_sheet.py --price C:\\другая\\папка

Результат: C:\\catalog\\scripts\\photo_matching.xlsx
  Лист 1 "Фото"   — извлечённые фото (A: file_name, B: original_name, C: заполнить вручную)
  Лист 2 "Товары" — товары из прайсов (A: название, B: категория, C: группа)
"""

import os
import sys
import json
import glob
import logging
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

DEFAULT_PRICE_DIR = r"C:\price"
PHOTO_MAP_PATH = SCRIPT_DIR / "photo_map.json"
OUTPUT_PATH = SCRIPT_DIR / "photo_matching.xlsx"


def load_env() -> None:
    for search_dir in [PROJECT_ROOT, SCRIPT_DIR]:
        env_path = search_dir / ".env"
        if env_path.exists():
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, value = line.partition("=")
                    os.environ.setdefault(key.strip(), value.strip().strip("'\""))
            return


def load_photo_map() -> list[dict]:
    if not PHOTO_MAP_PATH.exists():
        log.error("photo_map.json не найден: %s", PHOTO_MAP_PATH)
        log.error("Сначала запустите: python extract_photos.py")
        sys.exit(1)
    with open(PHOTO_MAP_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Убираем записи без file_name (товары без фото)
    return [e for e in data if e.get("file_name")]


def parse_excel_products(price_dir: Path) -> list[dict]:
    """Считать все товары из Excel-прайсов (упрощённо, как в upload.py)."""
    xlsx_files = sorted(glob.glob(str(price_dir / "*.xlsx")))
    if not xlsx_files:
        log.error("В папке %s нет .xlsx файлов", price_dir)
        sys.exit(1)

    # Загрузить category_map для display_group
    category_map: dict = {}
    map_path = SCRIPT_DIR / "category_map.json"
    if map_path.exists():
        with open(map_path, "r", encoding="utf-8") as f:
            category_map = json.load(f)

    all_products: list[dict] = []
    seen_names: set[str] = set()

    for filepath in xlsx_files:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        current_category = "Без категории"

        for row in ws.iter_rows(min_row=1, values_only=True):
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            c = row[2] if len(row) > 2 else None

            if a is None and b is None and c is None:
                continue
            b_str = str(b).strip().lower() if b else ""
            if b_str in ("цена",):
                continue
            if a is not None and str(a).strip() != "" and b is None and c is None:
                raw = str(a).strip()
                # убрать префикс 'а' у 'аКока-Кола'
                if len(raw) >= 2 and raw[0] == "а" and raw[1].isupper():
                    raw = raw[1:]
                current_category = raw
                continue
            if a is not None and b is not None:
                name = str(a).strip()
                if not name or name in seen_names:
                    continue
                seen_names.add(name)
                group = category_map.get(current_category, "Другое")
                all_products.append({
                    "name": name,
                    "category": current_category,
                    "group": group,
                })
        wb.close()

    log.info("Товаров из прайсов: %d (уникальных)", len(all_products))
    return all_products


def style_header_cell(cell, bg_hex: str = "4472C4") -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, bottom=thin)


def style_data_cell(cell, wrap: bool = False) -> None:
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.font = Font(size=10)


def set_column_width(ws, col_letter: str, width: float) -> None:
    ws.column_dimensions[col_letter].width = width


def build_sheet_photos(wb: openpyxl.Workbook, photos: list[dict]) -> None:
    ws = wb.create_sheet("Фото")

    # Заголовки
    headers = ["Файл фото", "Название из PDF", "Название из прайса (заполнить)"]
    bg_colors = ["2E75B6", "2E75B6", "375623"]
    for col, (header, color) in enumerate(zip(headers, bg_colors), 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell, color)

    ws.row_dimensions[1].height = 20

    # Данные
    for row_idx, entry in enumerate(photos, 2):
        ws.cell(row=row_idx, column=1, value=entry.get("file_name", ""))
        ws.cell(row=row_idx, column=2, value=entry.get("original_name", ""))
        # Колонка C — пустая, пользователь заполняет вручную
        ws.cell(row=row_idx, column=3, value="")

        for col in range(1, 4):
            style_data_cell(ws.cell(row=row_idx, column=col))

        # Лёгкая полосатость для читаемости
        if row_idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

    # Ширина колонок
    set_column_width(ws, "A", 38)
    set_column_width(ws, "B", 42)
    set_column_width(ws, "C", 46)

    # Заморозить первую строку
    ws.freeze_panes = "A2"

    log.info("Лист 'Фото': %d строк", len(photos))


def build_sheet_products(wb: openpyxl.Workbook, products: list[dict]) -> None:
    ws = wb.create_sheet("Товары")

    headers = ["Название товара", "Категория", "Группа"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell, "7030A0")

    ws.row_dimensions[1].height = 20

    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p["name"])
        ws.cell(row=row_idx, column=2, value=p["category"])
        ws.cell(row=row_idx, column=3, value=p["group"])

        for col in range(1, 4):
            style_data_cell(ws.cell(row=row_idx, column=col))

        if row_idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="F3EEFB")

    set_column_width(ws, "A", 52)
    set_column_width(ws, "B", 30)
    set_column_width(ws, "C", 24)

    ws.freeze_panes = "A2"

    # Автофильтр для удобного поиска
    ws.auto_filter.ref = f"A1:C{len(products) + 1}"

    log.info("Лист 'Товары': %d строк", len(products))


def main() -> None:
    load_env()

    parser = argparse.ArgumentParser(
        description="Создание Excel-файла для ручного сопоставления фото с товарами"
    )
    parser.add_argument(
        "--price",
        default=os.environ.get("EXCEL_DIR", DEFAULT_PRICE_DIR),
        help=f"Папка с .xlsx прайсами (по умолчанию: {DEFAULT_PRICE_DIR})",
    )
    args = parser.parse_args()

    price_dir = Path(args.price)
    if not price_dir.exists():
        log.error("Папка с прайсами не найдена: %s", price_dir)
        sys.exit(1)

    photos = load_photo_map()
    log.info("Фото из photo_map.json: %d", len(photos))

    products = parse_excel_products(price_dir)

    # Создать книгу (удалить дефолтный лист)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_sheet_photos(wb, photos)
    build_sheet_products(wb, products)

    wb.save(OUTPUT_PATH)
    log.info("Файл сохранён: %s", OUTPUT_PATH)

    print(f"\nГотово!")
    print(f"  Фото для сопоставления: {len(photos)}")
    print(f"  Товаров в прайсах:       {len(products)}")
    print(f"  Файл: {OUTPUT_PATH}")
    print()
    print("Инструкция:")
    print("  1. Откройте photo_matching.xlsx")
    print("  2. На листе 'Фото' заполните колонку C — скопируйте нужное название")
    print("     из листа 'Товары' (колонка A)")
    print("  3. Запустите: python apply_photo_matching.py")


if __name__ == "__main__":
    main()
