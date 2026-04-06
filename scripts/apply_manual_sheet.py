"""
apply_manual_sheet.py — Применение ручной привязки из photo_manual.xlsx

Читает photo_manual.xlsx (лист «Товары»):
  A: Название товара
  C: Файл фото → обновляет photo_overrides.json
  D: Описание   → обновляет description_overrides.json

Использование:
    python apply_manual_sheet.py
    python apply_manual_sheet.py --dry-run   # показать без сохранения
"""

import sys
import json
import logging
import argparse
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

SHEET_PATH = PROJECT_ROOT / "photo_manual.xlsx"
OVERRIDES_PATH = SCRIPT_DIR / "photo_overrides.json"
DESC_OVERRIDES_PATH = SCRIPT_DIR / "description_overrides.json"
PHOTO_URLS_PATH = SCRIPT_DIR / "photo_urls.json"

# Маппинг: Cloudinary папка → логическое имя (должно совпадать с upload.py)
CLOUDINARY_FOLDER_ALIAS: dict[str, str] = {"catalog": "akkond"}
# Известные логические папки (от пути на диске)
KNOWN_LOGICAL_FOLDERS = {"akkond", "presenter"}


def load_existing(path: Path) -> dict:
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_json(path: Path, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def build_url_index() -> dict[str, str]:
    """Строит индекс {logical_folder/filename → url} из photo_urls.json.

    Используется для определения логической папки файла по URL в Cloudinary.
    """
    import re as _re
    if not PHOTO_URLS_PATH.exists():
        return {}
    with open(PHOTO_URLS_PATH, "r", encoding="utf-8") as f:
        photo_urls: dict[str, str] = json.load(f)
    index: dict[str, str] = {}
    for filename, url in photo_urls.items():
        m = _re.search(r"/upload/v\d+/([^/]+)/", url)
        cloudinary_folder = m.group(1) if m else None
        logical_folder = CLOUDINARY_FOLDER_ALIAS.get(cloudinary_folder, cloudinary_folder)
        if logical_folder:
            index[f"{logical_folder}/{filename}"] = url
        index[filename] = url
    return index


def resolve_photo_key(raw_file: str, url_index: dict[str, str]) -> str:
    """Извлечь логический ключ «папка/имя» из пути или имени файла.

    Логика:
    1. Убрать file:/// префикс → получить путь на диске.
    2. Если родительский каталог — известная логическая папка (akkond/presenter) →
       вернуть «родитель/имя».
    3. Иначе искать имя в url_index → найдём ключ «folder/имя» → вернуть его.
    4. Не найдено → вернуть просто имя (файл не загружен в Cloudinary).
    """
    # 1. Убрать file:/// prefix и нормализовать
    clean = raw_file.replace("file:///", "").replace("file://", "")
    # На Windows /C:/path → C:/path
    if clean.startswith("/") and len(clean) > 2 and clean[2] == ":":
        clean = clean[1:]
    p = Path(clean)
    filename = p.name

    if not filename:
        return ""

    # Исправить двойное расширение: 869.jpgjpg → 869.jpg
    for ext in (".jpgjpg", ".jpegjpeg", ".pngpng", ".webpwebp"):
        if filename.lower().endswith(ext):
            trim = len(ext) // 2
            filename = filename[: len(filename) - trim]
            log.warning("Исправлено двойное расширение: %s → %s", raw_file, filename)
            break

    # 2. Если родитель — известная логическая папка
    parent = p.parent.name
    if parent in KNOWN_LOGICAL_FOLDERS:
        return f"{parent}/{filename}"

    # 3. Искать в url_index — найти ключ с папкой.
    # Файлы из корня photos/ (не из именованной подпапки) — presenter-товары,
    # поэтому предпочитаем presenter/ перед akkond/.
    matches = [key for key in url_index if "/" in key and key.split("/", 1)[1] == filename]
    for preferred in ("presenter", "akkond"):
        for key in matches:
            if key.split("/")[0] == preferred:
                return key
    if matches:
        return matches[0]

    # 4. Не найдено в Cloudinary
    return filename


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Применение ручной привязки фото и описаний из photo_manual.xlsx"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Показать результат без сохранения файлов",
    )
    args = parser.parse_args()

    if not SHEET_PATH.exists():
        log.error("Файл не найден: %s", SHEET_PATH)
        log.error("Сначала запустите: python scripts/make_manual_sheet.py")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True, read_only=True)

    if "Товары" not in wb.sheetnames:
        log.error("Лист «Товары» не найден в %s", SHEET_PATH)
        raise SystemExit(1)

    ws = wb["Товары"]

    # Строим индекс URL для резолвинга папок
    url_index = build_url_index()
    log.info("Загружено URL в индексе: %d", len(url_index))

    new_photos: dict[str, str] = {}    # {product_name: "folder/file_name"}
    new_descs: dict[str, str] = {}     # {product_name: description}
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        raw_file = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        description = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if not name:
            skipped += 1
            continue

        if raw_file:
            file_key = resolve_photo_key(raw_file, url_index)
            if not file_key:
                log.warning("Не удалось извлечь имя файла из: %s", raw_file)
                skipped += 1
                continue
            if name in new_photos:
                log.warning("Дубль товара «%s» — оставляю первое вхождение", name)
            else:
                new_photos[name] = file_key

        if description:
            if name not in new_descs:
                new_descs[name] = description

    wb.close()

    log.info(
        "Прочитано: фото=%d, описаний=%d, пустых строк=%d",
        len(new_photos), len(new_descs), skipped,
    )

    if not new_photos and not new_descs:
        print("Нет заполненных строк. Заполните колонки C и/или D в photo_manual.xlsx.")
        return

    if args.dry_run:
        if new_photos:
            print(f"\n--- Привязки фото (dry-run, {len(new_photos)} шт.) ---")
            for name, fn in list(new_photos.items())[:20]:
                print(f"  {name!r}  →  {fn}")
            if len(new_photos) > 20:
                print(f"  ... и ещё {len(new_photos) - 20}")
        if new_descs:
            print(f"\n--- Описания (dry-run, {len(new_descs)} шт.) ---")
            for name, desc in list(new_descs.items())[:10]:
                print(f"  {name!r}  →  {desc[:60]}...")
            if len(new_descs) > 10:
                print(f"  ... и ещё {len(new_descs) - 10}")
        return

    # ── Обновить photo_overrides.json ─────────────────────────────────────────
    photo_added = photo_updated = 0
    if new_photos:
        existing_photos = load_existing(OVERRIDES_PATH)
        log.info("Существующих привязок в photo_overrides.json: %d", len(existing_photos))
        for name, fn in new_photos.items():
            if name in existing_photos:
                photo_updated += 1
            else:
                photo_added += 1
        merged_photos = {**existing_photos, **new_photos}
        save_json(OVERRIDES_PATH, merged_photos)
        log.info("Сохранено в %s", OVERRIDES_PATH)

    # ── Обновить description_overrides.json ───────────────────────────────────
    desc_added = desc_updated = 0
    if new_descs:
        existing_descs = load_existing(DESC_OVERRIDES_PATH)
        log.info("Существующих описаний в description_overrides.json: %d", len(existing_descs))
        for name in new_descs:
            if name in existing_descs:
                desc_updated += 1
            else:
                desc_added += 1
        merged_descs = {**existing_descs, **new_descs}
        save_json(DESC_OVERRIDES_PATH, merged_descs)
        log.info("Сохранено в %s", DESC_OVERRIDES_PATH)

    # ── Итог ──────────────────────────────────────────────────────────────────
    print(f"\nПривязано фото:    {len(new_photos)}  (новых: {photo_added}, обновлено: {photo_updated})")
    print(f"Добавлено описаний: {len(new_descs)}  (новых: {desc_added}, обновлено: {desc_updated})")
    print()
    if new_photos:
        print(f"  → {OVERRIDES_PATH}")
    if new_descs:
        print(f"  → {DESC_OVERRIDES_PATH}")
    print()
    print("Следующий шаг: python scripts/upload.py  (обновит Google Sheet)")


if __name__ == "__main__":
    main()
