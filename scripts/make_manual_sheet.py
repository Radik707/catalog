"""
make_manual_sheet.py — Создание / дополнение Excel для ручной привязки фото и описаний

Режимы работы:
  Файл НЕ существует  → создать с нуля (все товары из прайсов).
  Файл существует     → ДОПОЛНИТЬ: добавить только новые товары, сохранив все
                        существующие строки без изменений.

Использование:
    python make_manual_sheet.py                       # авто-режим
    python make_manual_sheet.py --dry-run             # показать что будет добавлено, без записи
    python make_manual_sheet.py --rebuild             # пересоздать файл с нуля
    python make_manual_sheet.py --price C:\\другая\\папка

Результат: C:\\catalog\\photo_manual.xlsx
  Лист "Товары": A=Название, B=Группа, C=Файл фото, D=Описание, E=Статус
"""

import os
import sys
import json
import glob
import logging
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

DEFAULT_PRICE_DIR = r"C:\price"
OUTPUT_PATH = PROJECT_ROOT / "photo_manual.xlsx"


# ── Загрузка вспомогательных данных ───────────────────────────────────────────

def load_env() -> None:
    for search_dir in [PROJECT_ROOT, SCRIPT_DIR]:
        env_path = search_dir / ".env"
        if env_path.exists():
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, value = line.partition("=")
                    os.environ.setdefault(key.strip(), value.strip().strip("'\""))
            return


def load_category_map() -> dict:
    path = SCRIPT_DIR / "category_map.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


PRODUCT_OVERRIDES = {
    "Набор конфет ЛЮСИ": "Коробочные конфеты",
    "Набор конфет МВН": "Коробочные конфеты",
    "Набор конфет Сонуар": "Коробочные конфеты",
    "Фас. кор. конфеты": "Коробочные конфеты",
    "Вес. драже Арахис": "Детское",
    "Драже Скитлс": "Батончики и шоколад",
    "Драже М&М": "Батончики и шоколад",
    "Соус POMATO": "Соусы и специи",
    "POMATO": "Соусы и специи",
    "Соус": "Соусы и специи",
    "Ж/р Ментос": "Прикассовое",
    "Аджика АМЦА": "Соусы и специи",
    "Нап.кофейный": "Напитки",
    "Let's Be": "Напитки",
    "Киндер": "Батончики и шоколад",
    "Цикорий": "Прикассовое",
    "Холс": "Прикассовое",
}

PRODUCT_CONTAINS_OVERRIDES = [
    ("паста шок", "Батончики и шоколад"),
    ("шок. паста", "Батончики и шоколад"),
]


def get_group(name: str, source_category: str, category_map: dict) -> str:
    name_lower = name.lower()
    for prefix, group in PRODUCT_OVERRIDES.items():
        if name_lower.startswith(prefix.lower()):
            return group
    for substring, group in PRODUCT_CONTAINS_OVERRIDES:
        if substring.lower() in name_lower:
            return group
    return category_map.get(source_category, "Другое")


def parse_products(price_dir: Path, category_map: dict) -> list[dict]:
    """Загрузить уникальные товары из всех Excel-прайсов."""
    xlsx_files = sorted(glob.glob(str(price_dir / "*.xlsx")))
    if not xlsx_files:
        log.error("В папке %s нет .xlsx файлов", price_dir)
        sys.exit(1)

    seen: set[str] = set()
    products: list[dict] = []

    for filepath in xlsx_files:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        current_category = "Без категории"

        for row in ws.iter_rows(min_row=1, values_only=True):
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            c = row[2] if len(row) > 2 else None

            if a is None and b is None and c is None:
                continue

            b_str = str(b).strip().lower() if b else ""
            c_str = str(c).strip().lower() if c else ""
            if b_str in ("цена",) or c_str in ("остаток",):
                continue

            if a is not None and str(a).strip() and b is None and c is None:
                raw = str(a).strip()
                if len(raw) >= 2 and raw[0] == "а" and raw[1].isupper():
                    raw = raw[1:]
                current_category = raw
                continue

            if a is not None and b is not None:
                name = str(a).strip()
                if not name or name in seen:
                    continue
                try:
                    float(b)
                except (ValueError, TypeError):
                    continue
                seen.add(name)
                products.append({
                    "name": name,
                    "group": get_group(name, current_category, category_map),
                })

        wb.close()

    log.info("Товаров из прайсов: %d (уникальных)", len(products))
    return products


def load_photo_overrides() -> dict[str, str]:
    path = SCRIPT_DIR / "photo_overrides.json"
    if not path.exists():
        log.warning("photo_overrides.json не найден")
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    log.info("Загружено привязок фото: %d", len(data))
    return data


def load_file_to_description() -> dict[str, str]:
    path = SCRIPT_DIR / "photo_map.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        photo_map = json.load(f)
    result: dict[str, str] = {}
    for entry in photo_map:
        fn = entry.get("file_name")
        desc = (entry.get("description") or "").strip()
        if fn and desc and fn not in result:
            result[fn] = desc
    return result


def load_description_overrides() -> dict[str, str]:
    path = SCRIPT_DIR / "description_overrides.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ── Чтение существующего файла ─────────────────────────────────────────────────

def load_existing_names() -> set[str]:
    """Вернуть множество названий товаров из существующего photo_manual.xlsx."""
    if not OUTPUT_PATH.exists():
        return set()
    wb = openpyxl.load_workbook(OUTPUT_PATH, data_only=True, read_only=True)
    ws = wb.active
    names: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name and str(name).strip():
            names.add(str(name).strip())
    wb.close()
    log.info("Существующих товаров в файле: %d", len(names))
    return names


# ── Стили ──────────────────────────────────────────────────────────────────────

def style_header(cell, bg_hex: str) -> None:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, bottom=thin)


def style_data(cell, wrap: bool = False) -> None:
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    cell.font = Font(size=10)


# ── Режим: создать с нуля ──────────────────────────────────────────────────────

def build_sheet(
    wb: openpyxl.Workbook,
    products: list[dict],
    overrides: dict[str, str],
    file_to_desc: dict[str, str],
    desc_overrides: dict[str, str],
) -> tuple[int, int]:
    """Создать лист «Товары». Возвращает (с_фото, без_фото)."""
    ws = wb.create_sheet("Товары")

    headers = ["Название товара", "Группа", "Файл фото", "Описание товара", "Статус"]
    colors = ["2E75B6", "2E75B6", "375623", "7030A0", "C55A11"]
    for col, (header, color) in enumerate(zip(headers, colors), 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header(cell, color)
    ws.row_dimensions[1].height = 20

    rows_without: list[dict] = []
    rows_with: list[dict] = []

    for p in products:
        name = p["name"]
        file_name = overrides.get(name, "")
        description = desc_overrides.get(name, "")
        if not description and file_name:
            description = file_to_desc.get(file_name, "")

        row_data = {
            "name": name,
            "group": p["group"],
            "file_name": file_name,
            "description": description,
        }
        if file_name:
            rows_with.append(row_data)
        else:
            rows_without.append(row_data)

    rows_without.sort(key=lambda r: r["name"].lower())
    rows_with.sort(key=lambda r: r["name"].lower())
    all_rows = rows_without + rows_with

    for row_idx, r in enumerate(all_rows, 2):
        has_photo = bool(r["file_name"])
        bg = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
        if has_photo:
            bg = "EBF7EE" if row_idx % 2 == 0 else "F5FBF7"

        ws.cell(row=row_idx, column=1, value=r["name"])
        ws.cell(row=row_idx, column=2, value=r["group"])
        ws.cell(row=row_idx, column=3, value=r["file_name"])
        ws.cell(row=row_idx, column=4, value=r["description"])
        ws.cell(row=row_idx, column=5, value=f'=IF(C{row_idx}<>"","✅","❌")')

        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            style_data(cell, wrap=(col == 4))
            cell.fill = PatternFill("solid", fgColor=bg)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 10
    ws.freeze_panes = "A2"

    total = len(all_rows)
    ws.auto_filter.ref = f"A1:E{total + 1}"

    log.info("Лист «Товары»: %d строк", total)
    return len(rows_with), len(rows_without)


# ── Режим: дополнить существующий файл ────────────────────────────────────────

def append_new_products(new_products: list[dict]) -> None:
    """Дописать новые товары в конец существующего photo_manual.xlsx."""
    wb = openpyxl.load_workbook(OUTPUT_PATH)
    ws = wb.active

    start_row = ws.max_row + 1

    # Сгруппировать по категории, внутри — алфавит
    by_group: dict[str, list[str]] = defaultdict(list)
    for p in new_products:
        by_group[p["group"]].append(p["name"])
    for g in by_group:
        by_group[g].sort(key=str.lower)

    # Записывать группами в алфавитном порядке групп
    row_idx = start_row
    for group in sorted(by_group.keys()):
        for name in by_group[group]:
            bg = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"

            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=group)
            ws.cell(row=row_idx, column=3, value="")   # файл фото — заполнит вручную
            ws.cell(row=row_idx, column=4, value="")   # описание — заполнит вручную
            ws.cell(row=row_idx, column=5, value=f'=IF(C{row_idx}<>"","✅","❌")')

            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                style_data(cell, wrap=(col == 4))
                cell.fill = PatternFill("solid", fgColor=bg)

            row_idx += 1

    # Расширить автофильтр
    ws.auto_filter.ref = f"A1:E{row_idx - 1}"

    wb.save(OUTPUT_PATH)
    log.info("Файл обновлён: %s", OUTPUT_PATH)


# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    load_env()

    parser = argparse.ArgumentParser(
        description="Создание / дополнение Excel для ручной привязки фото и описаний"
    )
    parser.add_argument(
        "--price",
        default=os.environ.get("EXCEL_DIR", DEFAULT_PRICE_DIR),
        help=f"Папка с .xlsx прайсами (по умолчанию: {DEFAULT_PRICE_DIR})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Показать что будет добавлено, без изменения файла",
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="Пересоздать файл с нуля (игнорировать существующий)",
    )
    args = parser.parse_args()

    price_dir = Path(args.price)
    if not price_dir.exists():
        log.error("Папка с прайсами не найдена: %s", price_dir)
        sys.exit(1)

    category_map = load_category_map()
    all_products = parse_products(price_dir, category_map)

    file_exists = OUTPUT_PATH.exists() and not args.rebuild

    # ── Режим ДОПОЛНЕНИЯ ──────────────────────────────────────────────────────
    if file_exists:
        existing_names = load_existing_names()
        new_products = [p for p in all_products if p["name"] not in existing_names]

        if not new_products:
            print("\nНовых товаров не найдено — файл актуален.")
            return

        # Сгруппировать для отчёта
        by_group: dict[str, list[str]] = defaultdict(list)
        for p in new_products:
            by_group[p["group"]].append(p["name"])
        for g in by_group:
            by_group[g].sort(key=str.lower)

        print(f"\nСуществующих строк сохранено: {len(existing_names)}")
        print(f"Новых товаров для добавления:  {len(new_products)}")
        print()
        for group in sorted(by_group.keys()):
            items = by_group[group]
            print(f"  [{group}] ({len(items)}):")
            for name in items:
                print(f"    + {name}")

        if args.dry_run:
            print("\n[dry-run] Файл НЕ изменён.")
            return

        append_new_products(new_products)

        print(f"\nГотово!")
        print(f"  Сохранено старых строк: {len(existing_names)}")
        print(f"  Добавлено новых:        {len(new_products)}")
        print(f"  Файл:                   {OUTPUT_PATH}")

    # ── Режим СОЗДАНИЯ С НУЛЯ ─────────────────────────────────────────────────
    else:
        overrides = load_photo_overrides()
        file_to_desc = load_file_to_description()
        desc_overrides = load_description_overrides()

        if args.dry_run:
            by_group: dict[str, list[str]] = defaultdict(list)
            for p in all_products:
                by_group[p["group"]].append(p["name"])
            print(f"\nТоваров всего: {len(all_products)}")
            for group in sorted(by_group.keys()):
                print(f"  [{group}] ({len(by_group[group])})")
            print("\n[dry-run] Файл НЕ создан.")
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        with_photo, without_photo = build_sheet(
            wb, all_products, overrides, file_to_desc, desc_overrides
        )
        wb.save(OUTPUT_PATH)
        log.info("Файл создан: %s", OUTPUT_PATH)

        total = with_photo + without_photo
        print(f"\nГотово!")
        print(f"  Товаров всего:  {total}")
        print(f"  С фото:         {with_photo}")
        print(f"  Без фото:       {without_photo}")
        print(f"  Файл:           {OUTPUT_PATH}")

    print()
    print("Инструкция:")
    print("  1. Откройте photo_manual.xlsx")
    print("  2. Заполните колонку C (файл фото) и/или D (описание) для нужных товаров")
    print("  3. Запустите: python scripts/apply_manual_sheet.py")


if __name__ == "__main__":
    main()
