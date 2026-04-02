"""
apply_photo_matching.py — Применение ручного сопоставления фото → photo_overrides.json

Читает заполненный photo_matching.xlsx (лист "Фото", колонка C) и генерирует
photo_overrides.json: {"название из прайса": "file_name"}.

Использование:
    python apply_photo_matching.py
    python apply_photo_matching.py --dry-run   # показать без сохранения
"""

import json
import logging
import argparse
from pathlib import Path

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent

MATCHING_PATH = SCRIPT_DIR / "photo_matching.xlsx"
OVERRIDES_PATH = SCRIPT_DIR / "photo_overrides.json"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Применение ручного сопоставления фото с товарами"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Показать результат без сохранения photo_overrides.json",
    )
    args = parser.parse_args()

    if not MATCHING_PATH.exists():
        log.error("Файл не найден: %s", MATCHING_PATH)
        log.error("Сначала запустите: python make_photo_sheet.py")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(MATCHING_PATH, data_only=True, read_only=True)

    if "Фото" not in wb.sheetnames:
        log.error("Лист 'Фото' не найден в %s", MATCHING_PATH)
        raise SystemExit(1)

    ws = wb["Фото"]

    overrides: dict[str, str] = {}
    skipped_empty = 0
    skipped_bad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        file_name = str(row[0]).strip() if row[0] else ""
        product_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if not product_name:
            skipped_empty += 1
            continue

        if not file_name:
            log.warning("Пустой file_name при непустом названии '%s' — пропускаю", product_name)
            skipped_bad += 1
            continue

        if product_name in overrides:
            log.warning("Дублирующееся название '%s' — оставляю первое вхождение", product_name)
            continue

        overrides[product_name] = file_name

    wb.close()

    log.info(
        "Заполненных строк: %d | Пустых (без названия): %d | Пропущено из-за ошибок: %d",
        len(overrides),
        skipped_empty,
        skipped_bad,
    )

    if not overrides:
        print("Нет заполненных строк. Заполните колонку C на листе 'Фото'.")
        return

    if args.dry_run:
        print(f"\n--- Привязки (dry-run, {len(overrides)} шт.) ---")
        for name, fn in overrides.items():
            print(f"  {name!r}  →  {fn}")
        return

    # Загрузить существующие и объединить (новые перезаписывают старые)
    existing: dict[str, str] = {}
    if OVERRIDES_PATH.exists():
        with open(OVERRIDES_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)
        log.info("Существующих привязок в photo_overrides.json: %d", len(existing))

    merged = {**existing, **overrides}

    with open(OVERRIDES_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    added = len(overrides) - sum(1 for k in overrides if k in existing)
    updated = len(overrides) - added

    print(f"\nСоздано {len(merged)} привязок в photo_overrides.json")
    print(f"  Новых: {added} | Обновлено: {updated} | Итого в файле: {len(merged)}")
    print(f"  Файл: {OVERRIDES_PATH}")
    print()
    print("Следующий шаг: python upload.py  (пересоберёт Google Sheet с новыми привязками)")


if __name__ == "__main__":
    main()
