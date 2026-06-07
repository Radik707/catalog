"""
test_upload_new_format.py — Pytest-тесты парсера двух форматов Excel и логики новинок.

Покрывает:
- Новый формат 1С (название в D, цена/остаток в дальних колонках)
- Старый формат (A=категория/название, B=цена, C=остаток)
- Авто-определение формата (is_new_format, find_header_cols)
- Нормализацию названий (clean_new_name, normalize_name)
- Логику определения новинок (без вызова сети)

Все фикстуры .xlsx генерируются в памяти через openpyxl, бинарные файлы не коммитятся.
"""

import sys
from pathlib import Path
import pytest
import openpyxl

# Добавляем папку scripts в путь поиска модулей (для запуска из корня проекта)
SCRIPTS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPTS_DIR))

import upload  # noqa: E402  — импорт после sys.path


# ── Вспомогательные функции для создания xlsx-фикстур ──────────────────────

def _make_new_format_xlsx(tmp_path: Path) -> Path:
    """Создать xlsx-файл в новом формате 1С.

    Структура: строка-заголовок с «Цена» в колонке N (индекс 13) и «Остаток» в O (индекс 14);
    название в колонке D (индекс 3). Строки-дубли с названием в колонке A пропускаются.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Строка-заголовок: пустые ячейки до D и N/O
    header = [""] * 20
    header[3] = "Наименование"    # D
    header[13] = "Цена"           # N
    header[14] = "Остаток"        # O
    ws.append(header)

    # Товар 1: строка-дубль (название в A, D повторяет)
    row_dup = [""] * 20
    row_dup[0] = "Конфеты Ромашка"      # A — дубль, пропускаем
    row_dup[3] = "Конфеты Ромашка, шт"  # D — главная ячейка
    row_dup[13] = 120.50                 # N — цена
    row_dup[14] = 10                     # O — остаток
    ws.append(row_dup)

    # Товар 2: только D/N/O заполнены
    row2 = [""] * 20
    row2[3] = "Сок Добрый яблоко, л"
    row2[13] = 55.00
    row2[14] = 25
    ws.append(row2)

    # Строка-пустышка — должна быть пропущена
    ws.append([""] * 20)

    # Товар 3: нулевой остаток
    row3 = [""] * 20
    row3[3] = "Чай Greenfield, пак"
    row3[13] = 89.90
    row3[14] = 0
    ws.append(row3)

    path = tmp_path / "new_format.xlsx"
    wb.save(str(path))
    return path


def _make_old_format_xlsx(tmp_path: Path) -> Path:
    """Создать xlsx-файл в старом формате (A=название/категория, B=цена, C=остаток)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Строка-заголовок
    ws.append(["Наименование", "Цена", "Остаток"])

    # Строка-категория (A заполнена, B и C пустые)
    ws.append(["аНапитки", None, None])

    # Два товара в категории «Напитки»
    ws.append(["Вода Bon Aqua", 25.00, 50])
    ws.append(["Сок Rich яблоко", 65.50, 30])

    # Ещё одна категория
    ws.append(["аКонфеты", None, None])
    ws.append(["Мишка косолапый", 180.00, 15])

    path = tmp_path / "old_format.xlsx"
    wb.save(str(path))
    return path


def _make_old_format_xlsx_for_find_header(tmp_path: Path) -> Path:
    """Фикстура для find_header_cols: «Цена» в колонке B (индекс 1), «Остаток» в C (индекс 2)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Наименование", "Цена", "Остаток"])
    ws.append(["Товар А", 50.0, 10])
    path = tmp_path / "old_header.xlsx"
    wb.save(str(path))
    return path


# ── Тесты ─────────────────────────────────────────────────────────────────

class TestIsNewFormat:
    """Тесты автодетекции формата по индексу колонки «Цена»."""

    def test_new_format_col_13_returns_true(self):
        """price_col=13 (колонка N) → новый формат."""
        assert upload.is_new_format(13) is True

    def test_new_format_col_5_border_returns_true(self):
        """price_col=5 — граничное значение, при котором формат считается новым."""
        assert upload.is_new_format(5) is True

    def test_old_format_col_1_returns_false(self):
        """price_col=1 (колонка B) → старый формат."""
        assert upload.is_new_format(1) is False

    def test_old_format_col_4_returns_false(self):
        """price_col=4 — ещё не новый формат (порог ≥ 5)."""
        assert upload.is_new_format(4) is False

    def test_none_returns_false(self):
        """price_col=None (заголовок не найден) → не новый формат."""
        assert upload.is_new_format(None) is False


class TestFindHeaderCols:
    """Тесты поиска колонок «Цена» и «Остаток» в заголовке листа."""

    def test_new_format_finds_col_13_and_14(self, tmp_path):
        """В новом формате find_header_cols возвращает (13, 14)."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        price_col, stock_col = upload.find_header_cols(wb.active)
        wb.close()
        assert price_col == 13
        assert stock_col == 14

    def test_old_format_finds_col_1_and_2(self, tmp_path):
        """В старом формате find_header_cols возвращает (1, 2)."""
        xlsx_path = _make_old_format_xlsx_for_find_header(tmp_path)
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        price_col, stock_col = upload.find_header_cols(wb.active)
        wb.close()
        assert price_col == 1
        assert stock_col == 2

    def test_no_header_returns_none_none(self, tmp_path):
        """Если заголовок не найден — возвращается (None, None)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Просто", "данные", "без", "цены"])
        path = tmp_path / "no_header.xlsx"
        wb.save(str(path))

        wb2 = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        price_col, stock_col = upload.find_header_cols(wb2.active)
        wb2.close()
        assert price_col is None
        assert stock_col is None


class TestParseNewFormat:
    """Тесты парсера нового формата 1С."""

    def test_returns_correct_product_count(self, tmp_path):
        """parse_new_format возвращает только товарные строки (строка-заголовок исключена)."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        products = upload.parse_new_format(str(xlsx_path), price_col=13, stock_col=14)
        # 3 товарных строки: Конфеты Ромашка, Сок Добрый, Чай Greenfield
        assert len(products) == 3

    def test_name_cleaned_from_unit_suffix(self, tmp_path):
        """Хвостовая единица «, шт» убирается из названия через clean_new_name."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        products = upload.parse_new_format(str(xlsx_path), price_col=13, stock_col=14)
        names = [p["name"] for p in products]
        # «Конфеты Ромашка, шт» → «Конфеты Ромашка»
        assert "Конфеты Ромашка" in names
        assert "Конфеты Ромашка, шт" not in names

    def test_price_and_stock_correct(self, tmp_path):
        """Цена и остаток берутся из правильных колонок."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        products = upload.parse_new_format(str(xlsx_path), price_col=13, stock_col=14)
        romashka = next(p for p in products if "Ромашка" in p["name"])
        assert romashka["price"] == 120.50
        assert romashka["stock"] == 10

    def test_source_category_is_empty(self, tmp_path):
        """Все товары нового формата имеют source_category == '' (категорий нет)."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        products = upload.parse_new_format(str(xlsx_path), price_col=13, stock_col=14)
        for p in products:
            assert p["source_category"] == "", f"Ожидали пустую категорию, получили: {p['source_category']}"

    def test_duplicate_row_in_col_a_skipped(self, tmp_path):
        """Строка с названием только в A (строка-дубль) не попадает в результат как отдельный товар."""
        xlsx_path = _make_new_format_xlsx(tmp_path)
        products = upload.parse_new_format(str(xlsx_path), price_col=13, stock_col=14)
        # Товар «Конфеты Ромашка» должен быть ровно 1 раз
        matches = [p for p in products if "Ромашка" in p["name"]]
        assert len(matches) == 1, f"Ожидали 1 товар, найдено {len(matches)}: {matches}"


class TestParseOldFormat:
    """Тесты парсера старого формата (A=категория/название, B=цена, C=остаток)."""

    def test_products_get_source_category(self, tmp_path):
        """Товары получают source_category из ближайшей строки-категории."""
        xlsx_path = _make_old_format_xlsx(tmp_path)
        products = upload.parse_excel_file(str(xlsx_path))
        # Все товары должны иметь непустую source_category
        for p in products:
            assert p["source_category"] != "", f"Пустая категория у товара: {p['name']}"

    def test_category_prefix_stripped(self, tmp_path):
        """Префикс 'а' снимается у категорий-строк (аНапитки → Напитки)."""
        xlsx_path = _make_old_format_xlsx(tmp_path)
        products = upload.parse_excel_file(str(xlsx_path))
        categories = {p["source_category"] for p in products}
        # Категория должна быть без префикса 'а'
        assert "Напитки" in categories
        assert "аНапитки" not in categories

    def test_products_in_correct_category(self, tmp_path):
        """Товар 'Вода Bon Aqua' относится к категории 'Напитки', а не к другой."""
        xlsx_path = _make_old_format_xlsx(tmp_path)
        products = upload.parse_excel_file(str(xlsx_path))
        voda = next((p for p in products if "Вода Bon Aqua" in p["name"]), None)
        assert voda is not None
        assert voda["source_category"] == "Напитки"

    def test_price_and_stock_parsed(self, tmp_path):
        """Цена и остаток считываются правильно из старого формата."""
        xlsx_path = _make_old_format_xlsx(tmp_path)
        products = upload.parse_excel_file(str(xlsx_path))
        sok = next((p for p in products if "Rich" in p["name"]), None)
        assert sok is not None
        assert sok["price"] == 65.50
        assert sok["stock"] == 30

    def test_total_product_count(self, tmp_path):
        """Итоговое количество товаров совпадает с ожидаемым (3 товара в фикстуре)."""
        xlsx_path = _make_old_format_xlsx(tmp_path)
        products = upload.parse_excel_file(str(xlsx_path))
        assert len(products) == 3


class TestCleanAndNormalizeName:
    """Тесты нормализации названий (clean_new_name, normalize_name)."""

    def test_clean_removes_unit_suffix(self):
        """clean_new_name убирает хвостовую единицу после запятой."""
        assert upload.clean_new_name("Конфеты Ромашка, шт") == "Конфеты Ромашка"

    def test_clean_removes_kg_suffix(self):
        """clean_new_name убирает суффикс ', кг'."""
        assert upload.clean_new_name("Сахар, кг") == "Сахар"

    def test_clean_name_without_suffix_unchanged(self):
        """Название без суффикса-единицы остаётся без изменений."""
        assert upload.clean_new_name("Вода Bon Aqua") == "Вода Bon Aqua"

    def test_normalize_converts_to_lowercase(self):
        """normalize_name приводит к нижнему регистру."""
        assert upload.normalize_name("Конфеты Ромашка") == "конфеты ромашка"

    def test_normalize_removes_suffix_and_lowercases(self):
        """normalize_name убирает единицу и приводит к нижнему регистру."""
        assert upload.normalize_name("Конфеты Ромашка, шт") == "конфеты ромашка"

    def test_normalize_collapses_extra_spaces(self):
        """normalize_name схлопывает множественные пробелы."""
        result = upload.normalize_name("Конфеты  Ромашка,  шт")
        assert "  " not in result


class TestNovinkiLogic:
    """Тесты логики определения новинок (без вызова сети/Google Sheets).

    Воспроизводим поведение из main(): для товаров с source_category==""
    ищем normalized_name в current_groups и присваиваем группу или «Новинки».
    """

    def _apply_novinki(self, products: list[dict], current_groups: dict) -> tuple[list[dict], set]:
        """Воспроизвести ветку новинок из main() на синтетических данных."""
        new_names: set = set()
        for p in products:
            if p["source_category"] == "":
                g = current_groups.get(upload.normalize_name(p["name"]))
                if g and g != "Новинки":
                    p["display_group"] = g
                else:
                    p["display_group"] = "Новинки"
                    new_names.add(p["name"])
        return products, new_names

    def test_known_product_gets_group_from_catalog(self):
        """Известный товар (есть в current_groups) получает группу из каталога."""
        current_groups = {
            "конфеты ромашка": "Конфеты и печенье",
            "вода bon aqua": "Напитки",
        }
        products = [
            {"name": "Конфеты Ромашка", "source_category": "", "display_group": "Другое"},
        ]
        products, new_names = self._apply_novinki(products, current_groups)
        assert products[0]["display_group"] == "Конфеты и печенье"
        assert "Конфеты Ромашка" not in new_names

    def test_unknown_product_gets_novinki_group(self):
        """Неизвестный товар (нет в current_groups) получает группу «Новинки»."""
        current_groups = {
            "вода bon aqua": "Напитки",
        }
        products = [
            {"name": "Новый супер-товар", "source_category": "", "display_group": "Другое"},
        ]
        products, new_names = self._apply_novinki(products, current_groups)
        assert products[0]["display_group"] == "Новинки"
        assert "Новый супер-товар" in new_names

    def test_unknown_product_added_to_new_names(self):
        """Новый товар попадает в new_names (для бейджа «новинка»)."""
        current_groups: dict = {}
        products = [
            {"name": "Совсем новый товар", "source_category": "", "display_group": "Другое"},
        ]
        _, new_names = self._apply_novinki(products, current_groups)
        assert "Совсем новый товар" in new_names

    def test_old_format_product_not_affected(self):
        """Товар из старого формата (source_category != '') не попадает в Новинки."""
        current_groups: dict = {}
        products = [
            {"name": "Вода Bon Aqua", "source_category": "Напитки", "display_group": "Напитки"},
        ]
        products, new_names = self._apply_novinki(products, current_groups)
        # source_category непустой — ветка не выполняется
        assert products[0]["display_group"] == "Напитки"
        assert len(new_names) == 0

    def test_mixed_batch_only_new_format_goes_to_novinki(self):
        """В смешанной партии старые товары не трогаются, новые без группы → «Новинки»."""
        current_groups = {"вода bon aqua": "Напитки"}
        products = [
            # Старый формат: source_category непустой
            {"name": "Вода Bon Aqua", "source_category": "Напитки", "display_group": "Напитки"},
            # Новый формат: известный товар
            {"name": "Вода Bon Aqua", "source_category": "", "display_group": "Другое"},
            # Новый формат: совсем новый
            {"name": "Неизвестный товар", "source_category": "", "display_group": "Другое"},
        ]
        products, new_names = self._apply_novinki(products, current_groups)
        # Старый формат — без изменений
        assert products[0]["display_group"] == "Напитки"
        # Новый формат, известный → из каталога
        assert products[1]["display_group"] == "Напитки"
        # Новый формат, новый → Новинки
        assert products[2]["display_group"] == "Новинки"
        assert "Неизвестный товар" in new_names
        assert "Вода Bon Aqua" not in new_names
