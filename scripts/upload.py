r"""
upload.py — Скрипт-конвертер Excel → Google Sheet
Поддерживает два формата прайс-листов: старый (A=категория/название, B=цена, C=остаток)
и новый формат 1С (название в колонке D, «Цена»/«Остаток» в дальних колонках ≥ N).
Формат определяется автоматически по индексу колонки «Цена».

Новые товары (из нового формата, отсутствующие в текущем каталоге) получают
группу «Новинки» и бейдж «новинка». При недоступности текущих групп каталога
скрипт прерывается с кодом 1 (sys.exit), не допуская попадания всех товаров в «Новинки».

Режим --dry-run: парсинг без записи в Google Sheet. Используется для проверки
перед боевым прогоном (на сервере запускается через uploader/app.py с бэкапом через sheet_tool.py).

Использование:
    python upload.py                            # парсинг + запись в Google Sheet
    python upload.py --path /путь/к/папке       # указать папку с файлами
    python upload.py --dry-run                  # только парсинг, без записи

Настройка (.env в корне проекта):
    EXCEL_DIR=C:\price                          # папка с Excel-файлами (необязательно)
    GOOGLE_SHEETS_ID=ваш_id_таблицы             # ID Google Sheet
    GOOGLE_CREDENTIALS_PATH=credentials.json    # путь к ключу Service Account
"""

import os
import re
import sys
import json
import glob
import logging
import argparse
from pathlib import Path
from datetime import date

import openpyxl

# --- Кодировка вывода консоли ---
# Консоль Windows по умолчанию работает в cp1252 и не умеет печатать кириллицу
# через print()/логи — это роняло предпросмотр --dry-run с UnicodeEncodeError.
# Переключаем stdout/stderr на UTF-8 (по образцу фикса в sheet_helper.py).
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        # Поток не TextIOWrapper (перенаправлен/обёрнут) — пропускаем без ошибки
        pass

# --- Настройка логирования ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# --- Путь к папке скрипта (для category_map.json, .env) ---
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

# --- Путь к Excel-файлам по умолчанию ---
DEFAULT_EXCEL_DIR = r"C:\price"


def load_badges() -> dict:
    """Загрузить метки из badges.json. При отсутствии файла возвращает пустую структуру."""
    badges_path = SCRIPT_DIR / "badges.json"
    if not badges_path.exists():
        log.warning("Файл badges.json не найден: %s — метки не будут проставлены", badges_path)
        return {"исключения": [], "новинка": [], "хит": [], "акция": []}
    with open(badges_path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_badge(name: str, badges: dict) -> str:
    """Определить метку для товара по его названию.

    Приоритет: исключения → новинка → хит → акция.
    Поиск по частичному совпадению, регистронезависимый.
    """
    name_lower = name.lower()

    for exclusion in badges.get("исключения", []):
        if exclusion.lower() in name_lower:
            return ""

    for badge_key in ("новинка", "хит", "акция"):
        for substring in badges.get(badge_key, []):
            if substring.lower() in name_lower:
                return badge_key

    return ""


def load_category_map() -> dict:
    """Загрузить маппинг категорий из category_map.json."""
    map_path = SCRIPT_DIR / "category_map.json"
    if not map_path.exists():
        log.warning("Файл category_map.json не найден: %s", map_path)
        return {}
    with open(map_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_structure_map() -> dict:
    """Загрузить карту двухуровневой структуры из structure_map.json.

    Формат: { "Раздел": { "Подгруппа": ["КатегорияА", "КатегорияБ", ...] } }
    При отсутствии файла возвращает пустой словарь — скрипт работает без ошибок.
    """
    map_path = SCRIPT_DIR / "structure_map.json"
    if not map_path.exists():
        log.warning("Файл structure_map.json не найден: %s — поля Подгруппа/Раздел не будут заполнены", map_path)
        return {}
    with open(map_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_catalog_order() -> dict:
    """Загрузить пользовательский порядок витрины из вкладки «Настройки» (Работа 2).

    Владелец задаёт порядок разделов/подгрупп/категорий в админ-панели; панель пишет
    его JSON-строкой в ячейку «catalog_order» вкладки «Настройки». Формат JSON:
        {
          "sections":   ["Раздел A", "Раздел B", ...],
          "subgroups":  { "Раздел A": ["Подгруппа1", "Подгруппа2", ...] },
          "categories": { "Подгруппа1": ["КатА", "КатБ", ...] }
        }
    Любой уровень необязателен. Graceful-fallback: нет gspread/credentials/вкладки/ключа
    или битый JSON → {} (порядок из structure_map.json остаётся как есть).
    """
    try:
        import gspread  # noqa: F401
    except ImportError:
        return {}

    creds_path = os.environ.get("GOOGLE_CREDENTIALS_PATH", "")
    if not creds_path:
        for d in (SCRIPT_DIR, PROJECT_ROOT):
            c = d / "credentials.json"
            if c.exists():
                creds_path = str(c)
                break
    sheets_id = os.environ.get("GOOGLE_SHEETS_ID", "")
    if not creds_path or not sheets_id:
        return {}

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    try:
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        ss = gspread.authorize(creds).open_by_key(sheets_id)
        values = ss.worksheet("Настройки").get_all_values()
    except Exception as e:  # noqa: BLE001 — нет вкладки/сеть/доступ → порядок не задан
        log.info("Порядок каталога не прочитан (%s) — остаётся порядок structure_map.json", e)
        return {}

    raw = ""
    for row in values[1:]:  # пропускаем заголовок «Ключ | Значение»
        if row and row[0].strip() == "catalog_order":
            raw = row[1].strip() if len(row) > 1 else ""
            break
    if not raw:
        return {}
    try:
        order = json.loads(raw)
        return order if isinstance(order, dict) else {}
    except json.JSONDecodeError as e:
        log.warning("catalog_order: битый JSON (%s) — порядок игнорирован", e)
        return {}


def _reorder_by(names: list, desired: list) -> list:
    """Переставить names согласно desired; элементы вне desired — в конце в исходном порядке.

    Loss-free: возвращает ровно те же элементы, что и на входе (без дублей/потерь).
    Новые категории/подгруппы, которых владелец ещё не касался, оседают в конце уровня.
    """
    desired_set = set(desired)
    names_set = set(names)
    ordered = [n for n in desired if n in names_set]        # заданные владельцем, что реально есть
    rest = [n for n in names if n not in desired_set]       # новые/незаданные — в исходном порядке
    return ordered + rest


def apply_catalog_order(structure_map: dict, order: dict) -> dict:
    """Переупорядочить structure_map согласно сохранённому порядку владельца (Работа 2).

    Переставляет три уровня: разделы, подгруппы внутри раздела, категории внутри подгруппы.
    Ничего не теряет: элементы, отсутствующие в сохранённом порядке, встают в конец
    своего уровня в исходном порядке. Пустой order → карта возвращается без изменений.
    Downstream (build_category_index/apply_structure_mapping) читает порядок из позиции
    в dict — поэтому достаточно переставить ключи здесь, остальной код не меняется.
    """
    if not order:
        return structure_map
    sec_order = order.get("sections") or []
    sub_order = order.get("subgroups") or {}
    cat_order = order.get("categories") or {}

    new_map: dict = {}
    for section in _reorder_by(list(structure_map.keys()), sec_order):
        subgroups = structure_map[section]
        new_subs: dict = {}
        for subgroup in _reorder_by(list(subgroups.keys()), sub_order.get(section, [])):
            cats = list(subgroups[subgroup])
            new_subs[subgroup] = _reorder_by(cats, cat_order.get(subgroup, []))
        new_map[section] = new_subs
    return new_map


def build_category_index(structure_map: dict) -> dict:
    """Построить обратный индекс категорий для быстрого поиска раздела/подгруппы.

    Возвращает: { "НазваниеКатегории": (раздел, подгруппа, idx_раздела, idx_подгруппы, idx_кат) }
    Порядковые индексы используются для сортировки товаров по структуре (D-05).
    """
    index = {}
    for sec_idx, (section, subgroups) in enumerate(structure_map.items()):
        for sub_idx, (subgroup, categories) in enumerate(subgroups.items()):
            for cat_idx, category in enumerate(categories):
                index[category] = (section, subgroup, sec_idx, sub_idx, cat_idx)
    return index


def apply_structure_mapping(
    products: list[dict],
    category_index: dict,
) -> list[dict]:
    """Добавить поля display_subgroup и display_section из карты structure_map.json.

    Каждый товар получает:
      - p["display_subgroup"] — подгруппа (второй уровень навигации)
      - p["display_section"]  — раздел (первый уровень навигации)
      - p["_sort_key"]        — кортеж (idx_раздела, idx_подгруппы, idx_кат) для сортировки (D-05)

    Категории, не найденные в карте, попадают в fallback «Прочее»/«Прочее»
    с предупреждением в лог — ничего не теряется молча (D-12).
    Поле display_group (колонка «Группа») остаётся нетронутым (D-04).
    """
    FALLBACK_SECTION  = "Прочее"
    FALLBACK_SUBGROUP = "Прочее"
    FALLBACK_SORT     = (9999, 9999, 9999)

    unmapped = set()
    for product in products:
        # .get с fallback — мягкая деградация вместо KeyError, если у товара
        # вдруг нет поля source_category (соответствует обещанию D-12) (WR-02).
        cat = product.get("source_category", "")
        entry = category_index.get(cat)
        if entry is None:
            unmapped.add(cat)
            product["display_section"]  = FALLBACK_SECTION
            product["display_subgroup"] = FALLBACK_SUBGROUP
            product["_sort_key"]        = FALLBACK_SORT
        else:
            section, subgroup, sec_idx, sub_idx, cat_idx = entry
            product["display_section"]  = section
            product["display_subgroup"] = subgroup
            product["_sort_key"]        = (sec_idx, sub_idx, cat_idx)

    if unmapped:
        log.warning("Категории без структурного маппинга (попадут в «Прочее»):")
        for cat in sorted(unmapped):
            log.warning("  ! %s", cat)

    return products


def build_subgroup_index(structure_map: dict) -> dict:
    """Построить индекс подгрупп для разворота ручных правок типа «подгруппа» (этап 7).

    Возвращает: { "НазваниеПодгруппы": (раздел, sec_idx, sub_idx) }
    sec_idx/sub_idx — порядковые позиции раздела и подгруппы в карте (для сортировки).

    Если одна подгруппа встречается в нескольких разделах — берётся ПЕРВОЕ вхождение
    (последующие игнорируются), чтобы индекс был детерминирован.
    """
    index: dict = {}
    for sec_idx, (section, subgroups) in enumerate(structure_map.items()):
        for sub_idx, subgroup in enumerate(subgroups.keys()):
            if subgroup not in index:
                index[subgroup] = (section, sec_idx, sub_idx)
    return index


def apply_subgroup_overrides(products: list[dict], subgroup_index: dict) -> int:
    """Наложить ручные правки подгруппы владельца (тип «подгруппа», этап 7).

    Для каждого товара с p["subgroup_override"], значение которого ЕСТЬ в карте структуры,
    переписывает поля навигации:
      - p["display_subgroup"] = выбранная подгруппа
      - p["display_section"]  = раздел этой подгруппы (из карты)
      - p["_sort_key"]        = (sec_idx, sub_idx, 0) — товар встаёт в начало подгруппы

    Подгруппы из правки, которых НЕТ в карте, логируются предупреждением и пропускаются
    (товар остаётся с авто-маппингом из apply_structure_mapping). Возвращает число применённых правок.
    """
    applied = 0
    for p in products:
        override = p.get("subgroup_override")
        if not override:
            continue
        entry = subgroup_index.get(override)
        if entry is None:
            log.warning(
                "Ручная правка подгруппы «%s» для товара «%s» не найдена в structure_map.json — пропущена",
                override, p.get("name", ""),
            )
            continue
        section, sec_idx, sub_idx = entry
        p["display_subgroup"] = override
        p["display_section"] = section
        p["_sort_key"] = (sec_idx, sub_idx, 0)
        applied += 1
    return applied


def strip_category_prefix(name: str) -> str:
    """Убрать префикс 'а' у категорий вида 'аКока-Кола' → 'Кока-Кола'.

    Префикс удаляется только если:
    - строка начинается с 'а' (маленькая)
    - следующий символ — заглавная буква
    """
    if len(name) >= 2 and name[0] == "а" and name[1].isupper():
        return name[1:]
    return name


def is_header_row(a, b, c) -> bool:
    """Проверить, является ли строка заголовком (Цена/Остаток)."""
    b_str = str(b).strip().lower() if b else ""
    c_str = str(c).strip().lower() if c else ""
    return b_str in ("цена",) or c_str in ("остаток",)


def is_category_row(a, b, c) -> bool:
    """Категория = колонка A заполнена, B и C пустые."""
    return a is not None and str(a).strip() != "" and b is None and c is None


def parse_excel_file(filepath: str) -> list[dict]:
    """Распарсить один Excel-файл поставщика.

    Возвращает список словарей:
    [{name, price, stock, source_category, supplier_file}]
    """
    filename = Path(filepath).name
    log.info("Парсинг файла: %s", filename)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    products = []
    current_category = "Без категории"

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Извлекаем первые 3 колонки
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None

        # Пропускаем пустые строки
        if a is None and b is None and c is None:
            continue

        # Пропускаем строку-заголовок
        if is_header_row(a, b, c):
            continue

        # Строка-категория
        if is_category_row(a, b, c):
            raw_name = str(a).strip()
            current_category = strip_category_prefix(raw_name)
            continue

        # Строка-товар: A заполнена и B (цена) заполнена
        if a is not None and b is not None:
            name = str(a).strip()
            if not name:
                continue
            try:
                price = float(b)
            except (ValueError, TypeError):
                log.warning("  Не удалось прочитать цену: '%s' (строка: %s)", b, name)
                continue

            try:
                stock = int(float(c)) if c is not None else 0
            except (ValueError, TypeError):
                stock = 0

            products.append({
                "name": name,
                "price": round(price, 2),
                "stock": stock,
                "source_category": current_category,
                "supplier_file": filename,
            })

    wb.close()
    log.info("  Найдено товаров: %d", len(products))
    return products


# ── Поддержка НОВОГО формата выгрузки 1С ──
# В новом формате: название в колонке D (индекс 3), а «Цена»/«Остаток» — в дальних
# колонках (например N/O). Категорий-разделителей нет. Каждый товар продублирован
# строкой с названием в колонке A (двойная запятая) — её пропускаем.
NEW_FORMAT_NAME_COL = 3  # колонка D


def find_header_cols(ws) -> tuple[int | None, int | None]:
    """Найти индексы колонок «Цена» и «Остаток» в первых строках листа."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        price_col = stock_col = None
        for j, v in enumerate(row):
            if v is None:
                continue
            s = str(v).strip().lower()
            if s == "цена":
                price_col = j
            elif s == "остаток":
                stock_col = j
        if price_col is not None and stock_col is not None:
            return price_col, stock_col
    return None, None


def is_new_format(price_col: int | None) -> bool:
    """Новый формат — если «Цена» найдена в дальней колонке (D и правее), а не в B/C."""
    return price_col is not None and price_col >= 5


def clean_new_name(name: str) -> str:
    """Убрать хвостовую ', <ед.>' (', шт' / ', кг' / ', упак' и т.п.) из названия."""
    return re.sub(r",\s*[А-Яа-яA-Za-z.]+\s*$", "", name).strip()


def normalize_name(name: str) -> str:
    """Нормализовать название для сопоставления (без хвостовой единицы, нижний регистр)."""
    return re.sub(r"\s+", " ", clean_new_name(name)).strip().lower()


def parse_new_format(filepath: str, price_col: int, stock_col: int) -> list[dict]:
    """Распарсить файл нового формата. Категорий нет (source_category пустой)."""
    filename = Path(filepath).name
    log.info("Парсинг (новый формат): %s", filename)
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        name = row[NEW_FORMAT_NAME_COL] if len(row) > NEW_FORMAT_NAME_COL else None
        price = row[price_col] if len(row) > price_col else None
        stock = row[stock_col] if len(row) > stock_col else None
        if not name or price is None:
            continue
        name = str(name).strip()
        if name.lower() == "наименование":
            continue
        try:
            price = round(float(price), 2)
        except (ValueError, TypeError):
            continue
        try:
            stock = int(float(stock)) if stock is not None else 0
        except (ValueError, TypeError):
            stock = 0
        products.append({
            "name": clean_new_name(name),
            "price": price,
            "stock": stock,
            "source_category": "",   # категории нет — группа определяется по текущему каталогу
            "supplier_file": filename,
        })
    wb.close()
    log.info("  Найдено товаров: %d", len(products))
    return products


def load_current_groups() -> dict:
    """Прочитать текущий каталог (лист «Товары») → {норм_название: группа}.

    Нужно для нового формата: сохранить группировку существующих товаров и
    определить действительно новые (которых не было).
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return {}
    creds_path = os.environ.get("GOOGLE_CREDENTIALS_PATH", "")
    if not creds_path:
        for d in (SCRIPT_DIR, PROJECT_ROOT):
            c = d / "credentials.json"
            if c.exists():
                creds_path = str(c)
                break
    sheets_id = os.environ.get("GOOGLE_SHEETS_ID", "")
    if not creds_path or not sheets_id:
        return {}
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    try:
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        ss = gspread.authorize(creds).open_by_key(sheets_id)
        values = ss.worksheet("Товары").get_all_values()
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось прочитать текущие группы: %s", e)
        return {}
    if not values:
        return {}
    header = values[0]
    try:
        name_i = header.index("Наименование")
        grp_i = header.index("Группа")
    except ValueError:
        return {}
    mapping = {}
    for r in values[1:]:
        if len(r) > max(name_i, grp_i) and r[name_i] and r[grp_i]:
            mapping[normalize_name(r[name_i])] = r[grp_i]
    log.info("Загружено групп из текущего каталога: %d", len(mapping))
    return mapping


# ── Авто-новинка с истечением срока ──
# Товар, появившийся впервые, помечается «новинка» только NOVELTY_WINDOW_DAYS дней
# с даты первого появления, затем метка снимается автоматически. Дата хранится в
# scripts/novelty_dates.json (серверное runtime-состояние, НЕ в git): { норм_имя: 'YYYY-MM-DD' }.
NOVELTY_WINDOW_DAYS = 14   # сколько дней товар считается новинкой
NOVELTY_PRUNE_DAYS = 60    # старше этого — запись удаляется из файла (чистка)
NOVELTY_PATH = SCRIPT_DIR / "novelty_dates.json"


def load_novelty_dates() -> dict:
    """Загрузить даты первого появления { норм_имя: 'YYYY-MM-DD' }.

    Отсутствие файла / битый JSON → пустой dict (graceful): скрипт не падает,
    просто на этом прогоне ни у кого не будет авто-даты (метки выставит логика ниже).
    """
    try:
        with open(NOVELTY_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except FileNotFoundError:
        return {}
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось прочитать novelty_dates.json: %s — игнорирую", e)
        return {}


def save_novelty_dates(dates: dict) -> None:
    """Сохранить даты первого появления (атомарно через временный файл)."""
    try:
        tmp = NOVELTY_PATH.with_suffix(".json.tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(dates, f, ensure_ascii=False, indent=0)
        os.replace(tmp, NOVELTY_PATH)
    except Exception as e:  # noqa: BLE001
        log.warning("Не удалось сохранить novelty_dates.json: %s", e)


def compute_novelty(all_products: list, new_names: set, persist: bool) -> set:
    """Вычислить, какие товары сейчас «новинка» (дата первого появления ≤ 14 дней).

    - У товаров из new_names, которых ещё нет в файле дат, фиксируем дату = сегодня.
    - «Новинка» = есть дата И с неё прошло ≤ NOVELTY_WINDOW_DAYS дней.
    - Записи старше NOVELTY_PRUNE_DAYS удаляем (чистка файла).
    - persist=False (например, --dry-run) — состояние НЕ записывается на диск.

    Возвращает множество РАЗОВЫХ имён товаров (p["name"]), которым положена метка «новинка».
    """
    dates = load_novelty_dates()
    today = date.today()
    today_str = today.isoformat()

    # 1. Зафиксировать дату первого появления у новых товаров (если ещё не записана)
    for name in new_names:
        key = normalize_name(name)
        if key not in dates:
            dates[key] = today_str

    # 2. Определить актуальные новинки + почистить устаревшие записи
    fresh: set = set()
    pruned: dict = {}
    # Индекс присутствующих сейчас норм-имён — чтобы не хранить даты исчезнувших товаров вечно
    present_keys = {normalize_name(p["name"]) for p in all_products}
    for key, ds in dates.items():
        try:
            age = (today - date.fromisoformat(ds)).days
        except (ValueError, TypeError):
            continue  # битая дата — пропускаем (на следующем прогоне перезапишется)
        # Чистка: очень старые ИЛИ исчезнувшие из каталога записи не храним
        if age > NOVELTY_PRUNE_DAYS or key not in present_keys:
            continue
        pruned[key] = ds

    # 3. Сопоставить актуальные даты с товарами (по норм-имени) → метка «новинка»
    for p in all_products:
        ds = pruned.get(normalize_name(p["name"]))
        if not ds:
            continue
        try:
            age = (today - date.fromisoformat(ds)).days
        except (ValueError, TypeError):
            continue
        if age <= NOVELTY_WINDOW_DAYS:
            fresh.add(p["name"])

    if persist:
        save_novelty_dates(pruned)

    log.info(
        "Авто-новинка: актуальных новинок %d (окно %d дн.), записей в файле %d",
        len(fresh), NOVELTY_WINDOW_DAYS, len(pruned),
    )
    return fresh


# ── Память правок: надёжность чтения (страховка от тихой потери фото/правок) ──

class EditMemoryError(Exception):
    """Сбой чтения вкладки «Правки», который НЕЛЬЗЯ трактовать как «правок нет».

    Бросается, когда память правок не прочиталась из-за ошибки сети / лимита
    Google API / подозрительно усечённого ответа. Перегон обязан прерваться
    (sys.exit в main), а НЕ записать каталог без правок — иначе временный сбой
    тихо затрёт выставленные владельцем фото/группы/скрытия (баг «Добрый»).
    """


# Параметры чтения памяти правок
EDIT_MEMORY_RETRIES = 3                       # сколько раз пытаемся прочитать вкладку «Правки»
EDIT_MEMORY_RETRY_DELAYS = (1, 3, 6)          # паузы между попытками, сек (нарастающие)
EDIT_MEMORY_STATE_PATH = SCRIPT_DIR / "edit_memory_state.json"     # baseline числа правок
EDIT_MEMORY_ORPHANS_PATH = SCRIPT_DIR / "edit_memory_orphans.json"  # отчёт о несовпавших правках
EDIT_MEMORY_MIN_BASELINE = 50                 # ниже этого эталона baseline-проверку не делаем
EDIT_MEMORY_DROP_RATIO = 0.5                  # «подозрительно мало»: < 50% от прошлого успешного


def _load_edit_memory_baseline() -> "int | None":
    """Прочитать эталонное число правок из прошлого успешного чтения. None — если эталона нет."""
    try:
        with open(EDIT_MEMORY_STATE_PATH, "r", encoding="utf-8") as f:
            value = int(json.load(f).get("last_good_count"))
        return value if value >= 0 else None
    except Exception:  # noqa: BLE001 — нет файла/битый JSON → эталона просто нет
        return None


def _save_edit_memory_baseline(count: int) -> None:
    """Сохранить число правок как новый эталон (вызывается только после успешного чтения)."""
    try:
        with open(EDIT_MEMORY_STATE_PATH, "w", encoding="utf-8") as f:
            json.dump({"last_good_count": count}, f, ensure_ascii=False, indent=2)
    except Exception as e:  # noqa: BLE001 — не критично, эталон обновим в следующий раз
        log.warning("Не удалось сохранить эталон числа правок (%s) — продолжаю", e)


def _check_edit_memory_baseline(count: int) -> None:
    """Подушка для правок: резкое падение числа прочитанных правок = подозрение на сбой.

    Если в прошлый успешный раз правок было заметно больше, а сейчас < 50% —
    это похоже на частичный ответ Google API (без явного исключения).
    Прерываем перегон, чтобы не затереть выставленные фото/правки.
    """
    baseline = _load_edit_memory_baseline()
    if baseline is None or baseline < EDIT_MEMORY_MIN_BASELINE:
        return  # надёжного эталона нет — проверять не от чего (первый запуск/малая вкладка)
    if count < baseline * EDIT_MEMORY_DROP_RATIO:
        raise EditMemoryError(
            f"прочитано правок {count}, а в прошлый успешный раз — {baseline} "
            f"(меньше {int(EDIT_MEMORY_DROP_RATIO * 100)}%). Похоже на частичный ответ Google API. "
            "Перегон прерван, чтобы не затереть выставленные владельцем фото/правки. "
            f"Если правки реально удалены массово — удалите {EDIT_MEMORY_STATE_PATH.name} и повторите."
        )


def _read_pravki_values(creds_path: str, sheets_id: str, scopes: list) -> list:
    """Один заход чтения вкладки «Правки» → список строк (включая заголовок).

    WorksheetNotFound пробрасывается наружу как «вкладки нет» (легитимно пусто);
    остальные ошибки (лимит/сеть/доступ) уходят выше — там решают про ретрай.
    """
    import gspread  # noqa: F401 — нужен для типа исключения у вызывающего
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    ss = gspread.authorize(creds).open_by_key(sheets_id)
    return ss.worksheet("Правки").get_all_values()


def load_edit_memory() -> dict[str, dict[str, str]]:
    """Загрузить память ручных правок владельца из вкладки «Правки» той же Google-таблицы.

    Возвращает словарь {normalize_name(товар): {тип_правки: значение}},
    где тип правки — одно из: 'group', 'photo', 'description'.

    Схема вкладки «Правки» (по строке на правку):
      Колонка «Товар»    — нормализованное имя товара (ключ сопоставления, D-04)
      Колонка «Тип»      — тип правки: group | photo | description
      Колонка «Значение» — новое значение поля

    Пример строки: «конфеты ромашка» | «group» | «Коробочные конфеты»
    Несколько строк на один товар накапливаются в словарь типов.

    Различаем «правок реально нет» и «сбой чтения»:
      - нет gspread / нет credentials / нет вкладки (WorksheetNotFound) → {} (легитимно пусто);
      - ошибка сети / лимит Google API → ретрай EDIT_MEMORY_RETRIES раз, и при неудаче
        НЕ возвращаем {}, а бросаем EditMemoryError (перегон прервётся, фото/правки целы);
      - подозрительно малое число правок против эталона → тоже EditMemoryError.
    """
    import time

    try:
        import gspread
    except ImportError:
        return {}

    # --- Путь к credentials (те же переменные, что в load_current_groups) ---
    creds_path = os.environ.get("GOOGLE_CREDENTIALS_PATH", "")
    if not creds_path:
        for d in (SCRIPT_DIR, PROJECT_ROOT):
            c = d / "credentials.json"
            if c.exists():
                creds_path = str(c)
                break

    sheets_id = os.environ.get("GOOGLE_SHEETS_ID", "")
    if not creds_path or not sheets_id:
        return {}

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    # --- Чтение вкладки «Правки» с ретраями ---
    # Ошибку чтения (лимит/сеть) НЕ маскируем под «правок нет»: повторяем попытки,
    # и только WorksheetNotFound трактуем как легитимно пустую память.
    values = None
    last_err = None
    for attempt in range(1, EDIT_MEMORY_RETRIES + 1):
        try:
            values = _read_pravki_values(creds_path, sheets_id, scopes)
            break
        except gspread.exceptions.WorksheetNotFound:
            # Вкладки ещё нет — это нормально (память пуста), не ошибка
            log.info("Вкладка «Правки» не найдена — память правок пуста")
            return {}
        except Exception as e:  # noqa: BLE001 — лимит API / сеть / доступ → пробуем ещё раз
            last_err = e
            log.warning(
                "Чтение вкладки «Правки»: попытка %d/%d не удалась: %s",
                attempt, EDIT_MEMORY_RETRIES, e,
            )
            if attempt < EDIT_MEMORY_RETRIES:
                delay = EDIT_MEMORY_RETRY_DELAYS[min(attempt - 1, len(EDIT_MEMORY_RETRY_DELAYS) - 1)]
                time.sleep(delay)

    if values is None:
        # Все попытки провалились — это СБОЙ, а не «правок нет». Прерываем перегон.
        raise EditMemoryError(
            f"не удалось прочитать вкладку «Правки» за {EDIT_MEMORY_RETRIES} попыток: {last_err}"
        )

    if not values:
        # Вкладка существует, но пустая (даже без заголовка) — память реально пуста.
        # Эталон НЕ обновляем: baseline-проверка дальше не дойдёт (mapping будет пуст).
        return {}

    # --- Найти индексы колонок по заголовкам ---
    header = values[0]
    try:
        товар_i = header.index("Товар")
        тип_i = header.index("Тип")
        значение_i = header.index("Значение")
    except ValueError:
        # Лист прочитался, но заголовки не на месте — это повреждение схемы, а не «пусто».
        # Молча вернуть {} нельзя: затрёт правки. Прерываем перегон.
        raise EditMemoryError(
            "вкладка «Правки» прочитана, но не найдены колонки 'Товар'/'Тип'/'Значение' "
            f"(заголовок: {header[:8]}). Перегон прерван, чтобы не затереть правки."
        )

    # Допустимые типы правок (Этап 3: группа + фото + описание; Этап 4-02: название D-05; метка badge; Этап 7: подгруппа; Этап 8: скрыт)
    ALLOWED_TYPES = {"group", "photo", "description", "name", "badge", "подгруппа", "скрыт"}

    # --- Сборка словаря памяти ---
    mapping: dict[str, dict[str, str]] = {}
    valid_rows = 0  # принятых строк-правок (для диагностики частичного чтения, см. лог ниже)
    min_cols = max(товар_i, тип_i, значение_i) + 1
    for row in values[1:]:
        if len(row) < min_cols:
            continue
        # Значения ячеек трактуются строго как строки-данные (защита от инъекций, T-03-02)
        raw_product = str(row[товар_i]).strip()
        raw_type = str(row[тип_i]).strip()
        raw_value = str(row[значение_i]).strip()

        if not raw_product:
            log.warning("Вкладка «Правки»: пустой товар в строке — пропускаем")
            continue
        if not raw_type or raw_type not in ALLOWED_TYPES:
            log.warning(
                "Вкладка «Правки»: неизвестный тип правки '%s' для товара '%s' — пропускаем",
                raw_type, raw_product,
            )
            continue

        # Ключ — нормализованное имя товара (точное совпадение, D-04)
        key = normalize_name(raw_product)
        if key not in mapping:
            mapping[key] = {}
        mapping[key][raw_type] = raw_value
        valid_rows += 1

    # Подушка для правок: резкое падение числа против эталона → подозрение на сбой, прерываем.
    _check_edit_memory_baseline(len(mapping))
    # Чтение успешно и не подозрительно — обновляем эталон для следующих запусков.
    _save_edit_memory_baseline(len(mapping))

    # Логируем И товары, И строки: расхождение/проседание помогает заметить частичное чтение.
    log.info(
        "Загружено правок из памяти: %d товаров / %d строк (всего строк на листе: %d)",
        len(mapping), valid_rows, max(0, len(values) - 1),
    )
    return mapping


def apply_group_mapping(products: list[dict], category_map: dict) -> list[dict]:
    """Добавить поле display_group на основе маппинга категорий из category_map.json.

    Переопределения групп по конкретным товарам больше не хранятся в коде —
    они находятся во вкладке «Правки» (тип правки 'group') и применяются
    функцией apply_edit_memory() в main() после вызова этой функции (D-06).
    """
    unmapped = set()
    for product in products:
        cat = product["source_category"]
        group = category_map.get(cat)
        if group is None:
            unmapped.add(cat)
            group = "Другое"

        product["display_group"] = group

    if unmapped:
        log.warning("Категории без маппинга (попадут в 'Другое'):")
        for cat in sorted(unmapped):
            log.warning("  - %s", cat)

    return products


def apply_edit_memory(
    products: list[dict],
    edit_memory: dict[str, dict[str, str]],
) -> int:
    """Наложить память ручных правок владельца поверх авто-маппинга (D-06).

    Для каждого товара ищет правку по точному normalize_name(name) (D-04).
    Применяет только те типы, которые есть в записи памяти:
      - 'group'       → p['display_group'] = новое значение
      - 'photo'       → p['photo_override'] = URL (приоритет в products_to_rows)
      - 'description' → p['desc_override']  = текст (приоритет в products_to_rows)
      - 'name'        → p['display_name']   = отображаемое имя (ключ прайса не меняется, D-05)
      - 'badge'       → p['badge_override']  = метка («новинка»/«хит»/«акция»);
                        пустая строка означает явное снятие метки (приоритет в products_to_rows)

    Возвращает кортеж (число_товаров_без_правок, множество_ключей-сирот), где
    сироты — правки, чей ключ не совпал ни с одним товаром (вероятно, товар
    переименован в прайсе → ключ нормализованного имени «уехал», та же причина,
    по которой слетали фото «Добрый»). Их разбирает report_edit_memory_orphans().
    """
    new_for_memory = 0
    matched_keys: set = set()
    for p in products:
        key = normalize_name(p["name"])
        edit = edit_memory.get(key)
        if edit:
            matched_keys.add(key)
            # Правка группы — перебивает авто-маппинг (D-06)
            if "group" in edit:
                p["display_group"] = edit["group"]
            # Правки фото/описания — записываем в override-поля для точного совпадения (D-04)
            if "photo" in edit:
                p["photo_override"] = edit["photo"]
            if "description" in edit:
                p["desc_override"] = edit["description"]
            # Правка отображаемого названия (D-05, этап 4-02).
            # Ключ сопоставления остаётся normalize_name(p["name"]) из прайса — не меняется.
            # Только display_name используется при выводе в каталог; p["name"] — ключ прайса.
            if "name" in edit:
                p["display_name"] = edit["name"]
            # Правка метки (badge): ручное значение перебивает авто-определение.
            # Пустая строка — явное снятие метки («без метки»), её тоже сохраняем.
            if "badge" in edit:
                p["badge_override"] = edit["badge"]
            # Правка подгруппы (этап 7): владелец вручную переносит товар в подгруппу
            # двухуровневой структуры. Значение разворачивается в раздел/подгруппу/_sort_key
            # позже — в apply_subgroup_overrides() (после apply_structure_mapping).
            if "подгруппа" in edit:
                p["subgroup_override"] = edit["подгруппа"]
            # Правка скрытия (этап 8): «1» = скрыт с витрины, «» = показан.
            # Товар НЕ вырезается — остаётся в листе «Товары» с флагом (HIDE-02);
            # фильтрация происходит на стороне Next.js (lib/sheets.ts).
            if "скрыт" in edit:
                p["hidden"] = edit["скрыт"]
        else:
            new_for_memory += 1

    # Сироты: правки, чей ключ не совпал ни с одним товаром текущего прайса.
    orphan_keys = set(edit_memory) - matched_keys
    return new_for_memory, orphan_keys


def report_edit_memory_orphans(
    edit_memory: dict[str, dict[str, str]],
    orphan_keys: set,
) -> None:
    """Сообщить о правках, чей товар не найден в текущем прайсе (несовпадение ключа).

    Причина — товар переименован в прайсе ИЛИ временно отсутствует. Особенно
    болезненны сироты типа 'photo'/'name': владелец выставил фото/имя, товар
    переименовали → ключ перестал совпадать и правка тихо не применилась
    (механика бага «Добрый»). Громко логируем такие случаи и пишем полный
    отчёт в edit_memory_orphans.json (задел для показа владельцу в админ-панели).
    """
    if not orphan_keys:
        # Все правки нашли свои товары — снимаем устаревший отчёт, если был.
        try:
            if EDIT_MEMORY_ORPHANS_PATH.exists():
                EDIT_MEMORY_ORPHANS_PATH.unlink()
        except Exception:  # noqa: BLE001 — не критично
            pass
        return

    report = []
    photo_name = []  # подмножество с фото/именем — самые заметные для покупателя
    for key in sorted(orphan_keys):
        types = sorted(edit_memory.get(key, {}).keys())
        entry = {"key": key, "types": types}
        report.append(entry)
        if "photo" in types or "name" in types:
            photo_name.append(entry)

    log.warning(
        "Правок-сирот (товар не найден в прайсе — переименование или отсутствие): %d, "
        "из них с фото/именем: %d",
        len(report), len(photo_name),
    )
    for entry in photo_name[:30]:
        log.warning("  ! не применилось [%s]: %s", ", ".join(entry["types"]), entry["key"])
    if len(photo_name) > 30:
        log.warning("  … ещё %d правок с фото/именем (полный список в %s)",
                    len(photo_name) - 30, EDIT_MEMORY_ORPHANS_PATH.name)

    try:
        with open(EDIT_MEMORY_ORPHANS_PATH, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
    except Exception as e:  # noqa: BLE001 — отчёт необязателен
        log.warning("Не удалось записать отчёт сирот правок (%s)", e)


# Маппинг папок Cloudinary → логические имена в photo_overrides.json
CLOUDINARY_FOLDER_ALIAS: dict[str, str] = {"catalog": "akkond"}


def _build_url_index(photo_urls: dict[str, str]) -> dict[str, str]:
    """Строит индекс {logical_folder/filename → url} из photo_urls.json.

    Для каждого файла определяет папку Cloudinary по URL и применяет CLOUDINARY_FOLDER_ALIAS.
    Также добавляет записи без папки для обратной совместимости.
    """
    import re as _re
    index: dict[str, str] = {}
    for filename, url in photo_urls.items():
        m = _re.search(r"/upload/v\d+/([^/]+)/", url)
        cloudinary_folder = m.group(1) if m else None
        logical_folder = CLOUDINARY_FOLDER_ALIAS.get(cloudinary_folder, cloudinary_folder)
        if logical_folder:
            index[f"{logical_folder}/{filename}"] = url
        index[filename] = url  # обратная совместимость: ключ без папки
    return index


def load_url_index() -> dict[str, str]:
    """Загрузить индекс {folder/filename → url} и {filename → url} из photo_urls.json.

    Используется в products_to_rows() для резолва фото-référence из правки владельца
    (тип 'photo' хранит ссылку вида "presenter/файл.webp", а не полный URL Cloudinary).
    При отсутствии файла возвращает пустой словарь — скрипт работает без ошибок.
    """
    photo_urls_path = SCRIPT_DIR / "photo_urls.json"
    if not photo_urls_path.exists():
        return {}
    with open(photo_urls_path, "r", encoding="utf-8") as f:
        photo_urls: dict[str, str] = json.load(f)
    return _build_url_index(photo_urls)


def load_photo_data() -> dict[str, dict[str, str]]:
    """Загрузить маппинг название_товара → {url, description} из photo_map.json + photo_urls.json.

    Приоритет: photo_overrides.json (ручные привязки) → автоматический маппинг.
    photo_overrides.json поддерживает формат "folder/filename" и просто "filename".
    При отсутствии файлов возвращает пустой словарь — скрипт работает без ошибок.
    """
    photo_map_path = SCRIPT_DIR / "photo_map.json"
    photo_urls_path = SCRIPT_DIR / "photo_urls.json"
    overrides_path = SCRIPT_DIR / "photo_overrides.json"

    if not photo_map_path.exists() or not photo_urls_path.exists():
        if not photo_map_path.exists():
            log.warning("photo_map.json не найден — фото не будут добавлены")
        if not photo_urls_path.exists():
            log.warning("photo_urls.json не найден — фото не будут добавлены")
        return {}

    with open(photo_map_path, "r", encoding="utf-8") as f:
        photo_map = json.load(f)
    with open(photo_urls_path, "r", encoding="utf-8") as f:
        photo_urls: dict[str, str] = json.load(f)

    # Индекс: "folder/filename" и "filename" → url
    url_index = _build_url_index(photo_urls)

    # Автоматический маппинг: original_name (lower) → {url, description}
    name_to_data: dict[str, dict[str, str]] = {}
    for entry in photo_map:
        original_name = (entry.get("original_name") or "").strip()
        file_name = entry.get("file_name")
        description = (entry.get("description") or "").strip()
        if original_name and file_name and file_name in url_index:
            name_to_data[original_name.lower()] = {
                "url": url_index[file_name],
                "description": description,
            }

    # Ручные привязки перезаписывают авто-маппинг (description не перезаписывается)
    if overrides_path.exists():
        with open(overrides_path, "r", encoding="utf-8") as f:
            overrides: dict[str, str] = json.load(f)
        for product_name, file_key in overrides.items():
            if file_key in url_index:
                existing = name_to_data.get(product_name.lower(), {})
                name_to_data[product_name.lower()] = {
                    "url": url_index[file_key],
                    "description": existing.get("description", ""),
                }
        log.info("Загружено ручных привязок фото: %d", len(overrides))

    log.info("Загружено фото в маппинге: %d", len(name_to_data))
    return name_to_data


def _find_photo_entry(name: str, photo_data: dict[str, dict[str, str]]) -> dict[str, str] | None:
    """Найти запись о фото по частичному совпадению названия (регистронезависимо)."""
    if not photo_data:
        return None
    name_lower = name.lower()
    for photo_name, data in photo_data.items():
        if photo_name in name_lower or name_lower in photo_name:
            return data
    return None


def get_photo_url(name: str, photo_data: dict[str, dict[str, str]]) -> str:
    """Найти URL фото для товара по частичному совпадению (регистронезависимо)."""
    entry = _find_photo_entry(name, photo_data)
    return entry.get("url", "") if entry else ""


def get_photo_description(name: str, photo_data: dict[str, dict[str, str]]) -> str:
    """Найти описание товара по частичному совпадению (регистронезависимо)."""
    entry = _find_photo_entry(name, photo_data)
    return entry.get("description", "") if entry else ""


_CLOUD_NAME_CACHE: str | None = None


def _cloud_name(url_index: dict[str, str]) -> str:
    """Имя облака Cloudinary для сборки прямых ссылок.

    Берём из переменной CLOUDINARY_CLOUD_NAME, а если её нет в окружении (так бывает
    на сервере при запуске upload.py из загрузчика) — ВЫВОДИМ из любого готового URL
    в photo_urls.json (там все ссылки одного облака). Так резолв коротких ссылок
    («presenter/файл.jpg») не зависит от наличия переменной окружения. Результат кэшируется.
    """
    global _CLOUD_NAME_CACHE
    if _CLOUD_NAME_CACHE is not None:
        return _CLOUD_NAME_CACHE
    cloud = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
    if not cloud:
        for u in url_index.values():
            m = re.search(r"res\.cloudinary\.com/([^/]+)/", u)
            if m:
                cloud = m.group(1)
                break
    _CLOUD_NAME_CACHE = cloud
    return cloud


def _resolve_photo_override(ref: str, url_index: dict[str, str]) -> str:
    """Превратить фото-référence из правки владельца в полный URL Cloudinary.

    Правка типа 'photo' хранит ссылку вида "presenter/файл.webp" или "akkond_search/файл.jpg",
    а сайт ждёт полный https-URL. Логика резолва:
      - если значение уже начинается с http → вернуть как есть (это готовый URL);
      - иначе искать в url_index по значению целиком, затем по basename (имя файла);
      - если не нашлось, но ссылка содержит папку — собрать URL Cloudinary напрямую
        (панель грузит фото с public_id="presenter/<имя>", но НЕ пишет photo_urls.json,
        поэтому в url_index такого файла нет; URL Cloudinary детерминирован);
      - если совсем ничего → пустая строка (вызывающий код сделает fallback).
    """
    if not ref:
        return ""
    if ref.startswith("http"):
        return ref
    if ref in url_index:
        return url_index[ref]
    basename = ref.replace("\\", "/").split("/")[-1]
    if basename in url_index:
        return url_index[basename]
    # Последний резерв — реконструкция URL Cloudinary из ссылки-référence.
    # Так разворачиваются фото, загруженные через админ-панель (presenter/<имя>),
    # которых нет в photo_urls.json. Проверено: такой URL отдаёт картинку (200).
    cloud = _cloud_name(url_index)
    if cloud and "/" in ref:
        return f"https://res.cloudinary.com/{cloud}/image/upload/{ref}"
    return ""


def products_to_rows(
    products: list[dict],
    badges: dict | None = None,
    photo_data: dict[str, dict[str, str]] | None = None,
    new_names: set | None = None,
    url_index: dict[str, str] | None = None,
) -> list[list]:
    """Преобразовать список товаров в строки для Google Sheet.

    Формат: [Наименование, Цена, Остаток, Категория, Группа, Поставщик, Badge, ImageUrl, Description]
    Товары из new_names (реально новые) получают бейдж «новинка».

    Приоритет фото/описания (D-04, D-06):
      1. p['photo_override'] / p['desc_override'] — правка владельца (точное совпадение)
      2. get_photo_url / get_photo_description — авто-маппинг (частичное совпадение)

    photo_override хранит фото-référence ("presenter/файл.webp"), её резолвим в полный
    URL через url_index (см. _resolve_photo_override). Если в url_index не нашлось —
    fallback на авто-маппинг get_photo_url по названию товара.
    """
    if badges is None:
        badges = {"исключения": [], "новинка": [], "хит": [], "акция": []}
    if photo_data is None:
        photo_data = {}
    if url_index is None:
        url_index = {}
    new_names = new_names or set()
    header = [
        "Наименование", "Цена", "Остаток", "Категория",
        "Группа",           # старое поле — не трогаем (D-04, витрина переключится на этапе 6)
        "Поставщик", "Badge", "ImageUrl", "Description",
        "Подгруппа",        # новое поле этапа 5 (второй уровень навигации)
        "Раздел",           # новое поле этапа 5 (первый уровень навигации)
        "Скрыт",            # новое поле этапа 8 (флаг скрытия с витрины; «1» = скрыт)
    ]
    rows = [header]
    for p in products:
        # Приоритет: ручная правка метки владельца → авто-новинка → авто-метка из badges.json
        if "badge_override" in p:
            badge = p["badge_override"]        # "" = явно без метки; "хит"/"новинка" = ручная метка
        else:
            badge = "новинка" if p["name"] in new_names else get_badge(p["name"], badges)
        # Override-поля от apply_edit_memory имеют приоритет над авто-маппингом (D-04/D-06).
        # photo_override — это фото-référence ("presenter/файл.webp"), резолвим её в URL.
        # Если резолв не удался (нет в url_index) — fallback на авто-маппинг по названию.
        photo_ref = p.get("photo_override")
        if photo_ref:
            image_url = _resolve_photo_override(photo_ref, url_index) or get_photo_url(p["name"], photo_data)
        else:
            image_url = get_photo_url(p["name"], photo_data)
        description = p.get("desc_override") or get_photo_description(p["name"], photo_data)
        # Отображаемое название: display_name из правки (если есть), иначе имя из прайса (D-05).
        # get_badge / get_photo_url / get_photo_description по-прежнему получают p["name"] —
        # ключ из прайса не меняется, правка name затрагивает только вывод в каталог.
        displayed_name = p.get("display_name") or p["name"]
        rows.append([
            displayed_name,
            p["price"],
            p["stock"],
            p["source_category"],
            p["display_group"],                   # старое поле — нетронуто (D-04)
            p["supplier_file"],
            badge,
            image_url,
            description,
            p.get("display_subgroup", ""),        # новое поле — подгруппа (этап 5)
            p.get("display_section", ""),         # новое поле — раздел (этап 5)
            p.get("hidden", ""),                  # новое поле — скрытие (этап 8); «1» = скрыт
        ])
    return rows


def load_env() -> dict:
    """Загрузить переменные из .env файла (без python-dotenv).

    Ищет .env сначала в корне проекта, потом в папке скрипта.
    """
    env_vars = {}
    for search_dir in [PROJECT_ROOT, SCRIPT_DIR]:
        env_path = search_dir / ".env"
        if env_path.exists():
            log.info("Загружаю .env из: %s", env_path)
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, value = line.partition("=")
                    key = key.strip()
                    value = value.strip().strip("'\"")
                    env_vars[key] = value
                    os.environ.setdefault(key, value)
            break
    return env_vars


def upload_to_google_sheet(rows: list[list], num_files: int) -> None:
    """Записать данные в Google Sheet через gspread.

    1. Авторизация через Service Account
    2. Открыть таблицу по GOOGLE_SHEETS_ID
    3. Очистить лист "Товары" (или создать)
    4. Записать заголовки + товары
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        log.error(
            "Не установлены пакеты gspread / google-auth.\n"
            "Установите: pip install gspread google-auth"
        )
        sys.exit(1)

    # --- Путь к credentials ---
    creds_path = os.environ.get("GOOGLE_CREDENTIALS_PATH", "")
    if not creds_path:
        for search_dir in [SCRIPT_DIR, PROJECT_ROOT]:
            candidate = search_dir / "credentials.json"
            if candidate.exists():
                creds_path = str(candidate)
                break

    if not creds_path or not Path(creds_path).exists():
        log.error(
            "Файл credentials.json не найден.\n"
            "Укажите путь в .env: GOOGLE_CREDENTIALS_PATH=путь/к/credentials.json\n"
            "Или положите credentials.json в папку scripts/"
        )
        sys.exit(1)

    # --- ID таблицы ---
    sheets_id = os.environ.get("GOOGLE_SHEETS_ID", "")
    if not sheets_id:
        log.error(
            "Не указан GOOGLE_SHEETS_ID.\n"
            "Добавьте в .env: GOOGLE_SHEETS_ID=ваш_id_таблицы"
        )
        sys.exit(1)

    # --- Авторизация ---
    log.info("Авторизация в Google Sheets...")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc = gspread.authorize(credentials)

    # --- Открыть таблицу ---
    log.info("Открываю таблицу: %s", sheets_id)
    spreadsheet = gc.open_by_key(sheets_id)

    # --- Найти или создать лист "Товары" ---
    sheet_name = "Товары"
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
        log.info("Лист '%s' найден — очищаю...", sheet_name)
        worksheet.clear()
    except gspread.exceptions.WorksheetNotFound:
        log.info("Лист '%s' не найден — создаю...", sheet_name)
        # Ширину берём из заголовка: колонок стало 12 (добавлены «Подгруппа»/«Раздел»/«Скрыт»),
        # жёсткое cols=9 урезало бы новые столбцы на свежесозданном листе (WR-01).
        worksheet = spreadsheet.add_worksheet(
            title=sheet_name, rows=len(rows) + 10, cols=len(rows[0]) if rows else 12
        )

    # --- Записать данные пакетно ---
    log.info("Записываю %d строк (заголовок + %d товаров)...", len(rows), len(rows) - 1)

    if worksheet.row_count < len(rows):
        worksheet.resize(rows=len(rows) + 10)

    worksheet.update(rows, value_input_option="USER_ENTERED")

    num_products = len(rows) - 1
    log.info(
        "Загружено %d товаров из %d файлов в Google Sheet",
        num_products,
        num_files,
    )


def main():
    # Загрузить переменные окружения из .env
    load_env()

    parser = argparse.ArgumentParser(description="Конвертер Excel → Google Sheet")
    parser.add_argument(
        "--path",
        default=os.environ.get("EXCEL_DIR", DEFAULT_EXCEL_DIR),
        help="Папка с .xlsx файлами (по умолчанию: C:\\price\\)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только парсинг, без записи в Google Sheet",
    )
    args = parser.parse_args()

    excel_dir = Path(args.path)
    if not excel_dir.exists():
        log.error("Папка не найдена: %s", excel_dir)
        sys.exit(1)

    # Найти все .xlsx файлы
    xlsx_files = sorted(glob.glob(str(excel_dir / "*.xlsx")))
    if not xlsx_files:
        log.error("В папке %s нет .xlsx файлов", excel_dir)
        sys.exit(1)

    log.info("Найдено файлов: %d", len(xlsx_files))

    # Загрузить маппинг категорий
    category_map = load_category_map()
    log.info("Загружено категорий в маппинге: %d", len(category_map))

    # Загрузить метки
    badges = load_badges()
    badge_count = sum(len(v) for k, v in badges.items() if k != "исключения")
    log.info("Загружено меток: %d", badge_count)

    # Загрузить маппинг фото
    photo_data = load_photo_data()

    # Загрузить индекс URL для резолва фото-référence из правок владельца (тип 'photo')
    url_index = load_url_index()

    # Загрузить память ручных правок владельца (MEM-01).
    # При сбое чтения (лимит API / сеть / усечённый ответ) load_edit_memory бросает
    # EditMemoryError — прерываем перегон ДО записи, чтобы не затереть выставленные
    # фото/правки. Выход с кодом 1 → uploader/app.py откатит к Товары_BACKUP.
    try:
        edit_memory = load_edit_memory()
    except EditMemoryError as e:
        log.error("Перегон прерван — ненадёжное чтение памяти правок: %s", e)
        sys.exit(1)

    # Парсить все файлы (авто-определение формата: старый vs новый)
    all_products = []
    new_format_used = False
    for filepath in xlsx_files:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        price_col, stock_col = find_header_cols(wb.active)
        wb.close()
        if is_new_format(price_col):
            new_format_used = True
            all_products.extend(parse_new_format(filepath, price_col, stock_col))
        else:
            all_products.extend(parse_excel_file(filepath))

    log.info("Всего товаров из %d файлов: %d", len(xlsx_files), len(all_products))

    # Применить маппинг групп (старый формат — по category_map + переопределения)
    all_products = apply_group_mapping(all_products, category_map)

    # Новый формат: групп в файле нет — берём из текущего каталога по названию,
    # а товары, которых раньше не было, помечаем как новинки и кладём в «Новинки».
    new_names: set = set()
    if new_format_used:
        current_groups = load_current_groups()
        if not current_groups:
            log.error("Новый формат, но не удалось прочитать текущие группы каталога — "
                      "обновление прервано (иначе все товары попали бы в «Новинки»).")
            sys.exit(1)
        new_count = 0
        for p in all_products:
            if p["source_category"] == "":  # пришёл из нового формата
                g = current_groups.get(normalize_name(p["name"]))
                if g and g != "Новинки":
                    p["display_group"] = g
                else:
                    p["display_group"] = "Новинки"
                    new_names.add(p["name"])
                    new_count += 1
        log.info("Новых товаров (в «Новинки»): %d", new_count)

    # Наложить память правок поверх авто-маппинга (D-06: правка владельца побеждает)
    # Вызов ПОСЛЕ apply_group_mapping и блока нового формата, ПЕРЕД products_to_rows
    new_for_memory, orphan_keys = apply_edit_memory(all_products, edit_memory)
    log.info("Товаров без правок (новые для памяти): %d", new_for_memory)
    # Сообщить о правках, чей товар не нашёлся в прайсе (вероятно, переименован, п.3)
    report_edit_memory_orphans(edit_memory, orphan_keys)

    # [НОВОЕ, этап 5] Загрузить карту структуры и проставить Подгруппа/Раздел
    # Вызов ПОСЛЕ apply_edit_memory, ПЕРЕД products_to_rows (позиция из D-04/D-05)
    structure_map = load_structure_map()
    if structure_map:
        # [Работа 2] Переупорядочить разделы/подгруппы/категории по выбору владельца
        # (порядок из вкладки «Настройки»). Вызов ДО построения индексов — они берут
        # порядок из позиции ключей в карте. Пустой порядок → карта без изменений.
        catalog_order = load_catalog_order()
        if catalog_order:
            structure_map = apply_catalog_order(structure_map, catalog_order)
            log.info("Применён пользовательский порядок витрины (Работа 2)")
        category_index = build_category_index(structure_map)
        all_products = apply_structure_mapping(all_products, category_index)
        # [НОВОЕ, этап 7] Наложить ручные правки подгруппы поверх авто-маппинга по категории.
        # Вызов ПОСЛЕ apply_structure_mapping (правка владельца побеждает) и ПЕРЕД сортировкой —
        # чтобы переписанный _sort_key подхватился сортировкой ниже.
        subgroup_index = build_subgroup_index(structure_map)
        applied_subgroups = apply_subgroup_overrides(all_products, subgroup_index)
        log.info("применено ручных правок подгруппы: %d", applied_subgroups)
        # Сортировать товары по структуре: раздел → подгруппа → категория (D-05)
        # _sort_key проставлен apply_structure_mapping(); товары без структуры уходят в конец
        all_products.sort(key=lambda p: p.get("_sort_key", (9999, 9999, 9999)))
        log.info("Товары отсортированы по двухуровневой структуре (%d разделов)", len(structure_map))
        # D-12: сводка непокрытых — собраны, подсчитаны и залогированы
        uncovered = [p for p in all_products if p.get("display_section") == "Прочее"]
        if uncovered:
            uncovered_cats = sorted({p["source_category"] for p in uncovered})
            log.warning(
                "D-12: %d товаров (%d кат.) не покрыты structure_map.json → «Прочее»",
                len(uncovered), len(uncovered_cats),
            )
            for cat in uncovered_cats:
                log.warning("  ! %s", cat)
    else:
        log.info("structure_map.json не загружен — поля Подгруппа/Раздел будут пустыми")

    # Авто-новинка с истечением: метка «новинка» только в течение окна
    # NOVELTY_WINDOW_DAYS с даты первого появления. В --dry-run состояние не пишем.
    novelty_names = compute_novelty(all_products, new_names, persist=not args.dry_run)

    # Подготовить строки для Google Sheet (метку «новинка» получают актуальные новинки)
    rows = products_to_rows(all_products, badges, photo_data, novelty_names, url_index)

    # Статистика по группам
    groups = {}
    for p in all_products:
        g = p["display_group"]
        groups[g] = groups.get(g, 0) + 1
    log.info("Распределение по группам:")
    for g in sorted(groups, key=groups.get, reverse=True):
        log.info("  %s: %d товаров", g, groups[g])

    if args.dry_run:
        log.info("--dry-run: запись в Google Sheet пропущена")

        # D-13: человекочитаемый предпросмотр структуры (раздел → подгруппа → категории + счётчики)
        # Печатается ТОЛЬКО когда structure_map загружен, иначе — старый вывод первых 5 строк
        if structure_map:
            print("\n" + "=" * 70)
            print("ПРЕДПРОСМОТР ДВУХУРОВНЕВОЙ СТРУКТУРЫ (для сверки владельцем)")
            print("=" * 70)
            # Собрать счётчики товаров по (раздел, подгруппа)
            counters: dict = {}
            for p in all_products:
                key = (p.get("display_section", "Прочее"), p.get("display_subgroup", "Прочее"))
                counters[key] = counters.get(key, 0) + 1
            for section, subgroups in structure_map.items():
                total_section = sum(counters.get((section, sg), 0) for sg in subgroups)
                print(f"\n[{section}]  — итого: {total_section} товаров")
                for subgroup, categories in subgroups.items():
                    count = counters.get((section, subgroup), 0)
                    print(f"  {subgroup}  ({count} товаров)")
                    for cat in categories:
                        # Подсчитать товары этой категории среди реальных данных
                        cat_count = sum(
                            1 for p in all_products
                            if p["source_category"] == cat
                        )
                        if cat_count:
                            print(f"    · {cat}: {cat_count}")
            # Показать блок «Прочее» при наличии непокрытых категорий (D-12)
            uncovered_count = counters.get(("Прочее", "Прочее"), 0)
            if uncovered_count:
                print(f"\n[Прочее]  — {uncovered_count} товаров (не покрыты structure_map.json)")
                uncovered_cats = sorted({
                    p["source_category"] for p in all_products
                    if p.get("display_section") == "Прочее"
                })
                for cat in uncovered_cats:
                    cnt = sum(1 for p in all_products if p["source_category"] == cat)
                    print(f"    · {cat}: {cnt}")
            print("=" * 70)
            print(f"\nВсего товаров: {len(all_products)}")
        else:
            # Старый dry-run вывод — когда structure_map не загружен
            print("\nПример данных (первые 5 товаров):")
            print("-" * 100)
            for row in rows[:6]:  # заголовок + 5 товаров
                print(" | ".join(str(x) for x in row))
            print("-" * 100)
            print(f"Всего строк (без заголовка): {len(rows) - 1}")
    else:
        upload_to_google_sheet(rows, num_files=len(xlsx_files))

    return rows


if __name__ == "__main__":
    main()
